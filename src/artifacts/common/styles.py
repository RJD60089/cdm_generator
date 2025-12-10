# src/artifacts/common/styles.py
"""Shared styling constants for Excel and Word artifacts."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelStyles:
    """Excel formatting styles."""
    
    # Colors
    HEADER_BG = "4472C4"  # Blue
    HEADER_FG = "FFFFFF"  # White text
    ALT_ROW_BG = "D9E2F3"  # Light blue
    PK_BG = "FFF2CC"  # Light yellow for PKs
    FK_BG = "E2EFDA"  # Light green for FKs
    WARNING_BG = "FCE4D6"  # Light orange for review items
    
    # Fonts
    HEADER_FONT = Font(bold=True, color=HEADER_FG, size=11)
    BODY_FONT = Font(size=10)
    BOLD_FONT = Font(bold=True, size=10)
    LINK_FONT = Font(color="0563C1", underline="single", size=10)
    
    # Fills
    HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    ALT_FILL = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")
    PK_FILL = PatternFill(start_color=PK_BG, end_color=PK_BG, fill_type="solid")
    FK_FILL = PatternFill(start_color=FK_BG, end_color=FK_BG, fill_type="solid")
    WARNING_FILL = PatternFill(start_color=WARNING_BG, end_color=WARNING_BG, fill_type="solid")
    
    # Alignment
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    
    # Borders
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Column widths (reasonable defaults)
    DEFAULT_WIDTHS = {
        "entity": 25,
        "attribute": 30,
        "description": 50,
        "data_type": 15,
        "boolean": 10,
        "code": 20,
        "date": 15,
        "narrow": 8,
        "medium": 20,
        "wide": 40,
        "extra_wide": 60
    }
    
    @classmethod
    def apply_header_style(cls, cell):
        """Apply header styling to a cell."""
        cell.font = cls.HEADER_FONT
        cell.fill = cls.HEADER_FILL
        cell.alignment = cls.HEADER_ALIGN
        cell.border = cls.THIN_BORDER
    
    @classmethod
    def apply_body_style(cls, cell, is_alt_row: bool = False):
        """Apply body styling to a cell."""
        cell.font = cls.BODY_FONT
        cell.alignment = cls.BODY_ALIGN
        cell.border = cls.THIN_BORDER
        if is_alt_row:
            cell.fill = cls.ALT_FILL
    
    @classmethod
    def apply_pk_style(cls, cell):
        """Highlight primary key cells."""
        cell.fill = cls.PK_FILL
        cell.font = cls.BOLD_FONT
    
    @classmethod
    def apply_fk_style(cls, cell):
        """Highlight foreign key cells."""
        cell.fill = cls.FK_FILL
    
    @classmethod
    def set_column_widths(cls, sheet, width_map: dict):
        """Set column widths from a mapping of column letter to width."""
        for col, width in width_map.items():
            sheet.column_dimensions[col].width = width


class WordStyles:
    """Word document styling constants (for future use)."""
    
    HEADING_1_SIZE = 16
    HEADING_2_SIZE = 14
    HEADING_3_SIZE = 12
    BODY_SIZE = 11
    
    TABLE_HEADER_BG = "4472C4"
    TABLE_ALT_BG = "D9E2F3"
