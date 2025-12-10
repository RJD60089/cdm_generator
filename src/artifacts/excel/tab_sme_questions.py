# src/artifacts/excel/tab_sme_questions.py
"""SME Questions tab for Excel CDM."""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.gap_extractor import GapExtractor


def create_sme_questions_tab(
    wb: Workbook, 
    extractor: CDMExtractor,
    gap_extractor: GapExtractor = None
) -> Worksheet:
    """
    Create SME Questions tab.
    
    Shows questions for subject matter experts from gap analysis.
    """
    
    ws = wb.create_sheet("SME_Questions")
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # Headers
    headers = [
        "Question ID",
        "Category",
        "Related Entities", 
        "Question",
        "Context",
        "SME Response",
        "Resolution"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    row = 2
    
    if gap_extractor:
        questions = gap_extractor.get_sme_questions()
        
        for q in questions:
            ws.cell(row=row, column=1, value=q.question_id).border = thin_border
            ws.cell(row=row, column=2, value=q.category).border = thin_border
            
            # Related entities as comma-separated string
            related = ", ".join(q.related_entities) if q.related_entities else ""
            ws.cell(row=row, column=3, value=related).border = thin_border
            
            ws.cell(row=row, column=4, value=q.question_text).border = thin_border
            ws.cell(row=row, column=4).alignment = wrap_alignment
            ws.cell(row=row, column=5, value=q.context or "").border = thin_border
            ws.cell(row=row, column=5).alignment = wrap_alignment
            ws.cell(row=row, column=6, value="").border = thin_border  # For SME to fill
            ws.cell(row=row, column=7, value="").border = thin_border  # For resolution
            row += 1
    
    if row == 2:
        # No data
        ws.cell(row=2, column=1, value="No SME questions identified")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    
    # Column widths
    widths = [12, 15, 30, 50, 40, 30, 30]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    # Auto-filter
    if row > 2:
        ws.auto_filter.ref = f"A1:G{row-1}"
    
    return ws