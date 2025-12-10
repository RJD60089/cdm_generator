# src/artifacts/excel/tab_relationships.py
"""Generate Relationships tab for Excel CDM."""

from openpyxl import Workbook
from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.styles import ExcelStyles


def create_relationships_tab(wb: Workbook, extractor: CDMExtractor) -> None:
    """
    Create the Relationships tab with FK details.
    
    Columns:
    - Parent Entity, Parent Key, Child Entity, Foreign Key
    - Relationship Type, Description
    """
    
    ws = wb.create_sheet("Relationships")
    
    # Headers
    headers = [
        "Parent Entity", "Parent Key", 
        "Child Entity", "Foreign Key",
        "Relationship Type", "Description"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        ExcelStyles.apply_header_style(cell)
    
    # Data rows
    relationships = extractor.get_relationships()
    
    for row_idx, rel in enumerate(relationships, 2):
        is_alt = row_idx % 2 == 0
        
        row_data = [
            rel.parent_entity,
            rel.parent_key,
            rel.child_entity,
            rel.foreign_key,
            rel.relationship_type,
            rel.description or ""
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            ExcelStyles.apply_body_style(cell, is_alt)
    
    # Column widths
    widths = {
        "A": 25,  # Parent Entity
        "B": 25,  # Parent Key
        "C": 25,  # Child Entity
        "D": 25,  # Foreign Key
        "E": 18,  # Relationship Type
        "F": 50   # Description
    }
    ExcelStyles.set_column_widths(ws, widths)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Auto-filter
    if relationships:
        ws.auto_filter.ref = f"A1:F{len(relationships) + 1}"
