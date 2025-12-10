# src/artifacts/excel/tab_source_files.py
"""Source Files tab for Excel CDM."""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.config.config_parser import AppConfig


def create_source_files_tab(wb: Workbook, config: AppConfig) -> Worksheet:
    """
    Create Source Files tab.
    
    Shows input files used to generate the CDM.
    """
    
    ws = wb.create_sheet("Source_Files")
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, size=12)
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Configuration header
    ws.cell(row=row, column=1, value="Configuration").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    
    # Config values
    ws.cell(row=row, column=1, value="Domain").border = thin_border
    ws.cell(row=row, column=2, value=config.cdm.domain).border = thin_border
    row += 1
    ws.cell(row=row, column=1, value="Type").border = thin_border
    ws.cell(row=row, column=2, value=config.cdm.type).border = thin_border
    row += 1
    ws.cell(row=row, column=1, value="Output Directory").border = thin_border
    ws.cell(row=row, column=2, value=config.output.directory).border = thin_border
    row += 2
    
    # Guardrails files
    if config.has_guardrails():
        ws.cell(row=row, column=1, value="Guardrails Files").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        
        for gfile in config.guardrails:
            ws.cell(row=row, column=1, value=str(gfile)).border = thin_border
            row += 1
        row += 1
    
    # Glue files
    if config.has_glue():
        ws.cell(row=row, column=1, value="Glue Files").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        
        for glue_file in config.glue:
            ws.cell(row=row, column=1, value=str(glue_file)).border = thin_border
            row += 1
        row += 1
    
    # DDL files
    if config.ddl:
        ws.cell(row=row, column=1, value="DDL Files").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        
        for ddl_file in config.ddl:
            ws.cell(row=row, column=1, value=str(ddl_file)).border = thin_border
            row += 1
        row += 1
    
    # NCPDP standards
    if config.has_ncpdp():
        ws.cell(row=row, column=1, value="NCPDP Standards").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        
        # Headers
        ws.cell(row=row, column=1, value="Standard").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value="Version").font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=3, value="File").font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        row += 1
        
        for standard in config.ncpdp_general_standards + config.ncpdp_script_standards:
            ws.cell(row=row, column=1, value=standard.get('name', '-')).border = thin_border
            ws.cell(row=row, column=2, value=standard.get('version', '-')).border = thin_border
            ws.cell(row=row, column=3, value=standard.get('file', '-')).border = thin_border
            row += 1
        row += 1
    
    # FHIR resources
    if config.has_fhir():
        ws.cell(row=row, column=1, value="FHIR Resources").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        
        # Headers
        ws.cell(row=row, column=1, value="Resource Type").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value="File Type").font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=3, value="File").font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        row += 1
        
        for resource in config.fhir_igs:
            ws.cell(row=row, column=1, value=resource.get('resource_type', resource.get('name', '-'))).border = thin_border
            ws.cell(row=row, column=2, value=resource.get('file_type', '-')).border = thin_border
            ws.cell(row=row, column=3, value=resource.get('file', '-')).border = thin_border
            row += 1
        row += 1
    
    # Naming standards
    if config.naming_standard:
        ws.cell(row=row, column=1, value="Naming Standards").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        
        for ns_file in config.naming_standard:
            ws.cell(row=row, column=1, value=str(ns_file)).border = thin_border
            row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 60
    
    return ws