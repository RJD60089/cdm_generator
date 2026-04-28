# src/artifacts/excel/tab_cross_reference.py
"""
Cross-Reference Tab Generator

Shows source lineage for each attribute - maps CDM attributes back to 
original source fields (EDW, Guardrails, Glue, NCPDP, FHIR, etc.).

Source columns are built dynamically from whatever sources are present in
the Full CDM, so adding EDW (or any future source) requires no code changes.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.schema_resolver import (
    ancillary_attribute_index,
    format_ancillary_source_refs,
)


# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Preferred display order for known source types.
# Any unknown sources are appended alphabetically after these.
_SOURCE_ORDER = ["edw", "guardrails", "glue", "ncpdp", "fhir"]


def _format_source_label(source_type: str) -> str:
    """Convert source_type key to a readable column header."""
    upper_as_is = {"ncpdp", "fhir", "edw"}
    if source_type.lower() in upper_as_is:
        return f"{source_type.upper()} Source"
    return f"{source_type.title()} Source"


def _extract_source_field(lineage: Dict, source_type: str) -> str:
    """Extract source field name(s) from lineage data for a given source type."""
    source_data = lineage.get(source_type, [])

    if not source_data:
        return ""

    # Handle list of mappings
    if isinstance(source_data, list):
        fields = []
        for mapping in source_data:
            if isinstance(mapping, dict):
                entity = mapping.get("source_entity", "")
                attr = mapping.get("source_attribute", "")
                if entity and attr:
                    fields.append(f"{entity}.{attr}")
                elif attr:
                    fields.append(attr)
                elif entity:
                    fields.append(entity)
            elif isinstance(mapping, str):
                fields.append(mapping)
        return "; ".join(fields) if fields else ""

    # Handle dict directly
    if isinstance(source_data, dict):
        entity = source_data.get("source_entity", "")
        attr = source_data.get("source_attribute", "")
        if entity and attr:
            return f"{entity}.{attr}"
        return attr or entity or ""

    # Handle plain string
    if isinstance(source_data, str):
        return source_data

    return ""


def _discover_source_types(extractor: CDMExtractor) -> List[str]:
    """
    Walk all attribute source_lineage entries in the CDM and return
    the unique source types present, in preferred display order.
    """
    found = set()
    for attr in extractor.get_all_attributes():
        lineage = attr.source_lineage or {}
        found.update(lineage.keys())

    # Also check entity-level source_lineage if present
    for entity in extractor.get_entities():
        # EntitySummary is an object, not a dict - use getattr
        entity_lineage = getattr(entity, "source_lineage", {}) or {}
        found.update(entity_lineage.keys())

    # Sort: known order first, then any extras alphabetically
    ordered = [s for s in _SOURCE_ORDER if s in found]
    extras = sorted(s for s in found if s not in _SOURCE_ORDER)
    return ordered + extras


def create_cross_reference_tab(
    wb: Workbook,
    extractor: CDMExtractor,
    outdir: Optional[Path] = None,
    cdm_name: str = "",
) -> Worksheet:
    """
    Create Cross-Reference tab showing attribute-to-source mappings.

    Fixed columns:
        Entity | Attribute | Data Type

    Dynamic source columns (one per source type found in the CDM):
        EDW Source | Guardrails Source | Glue Source | NCPDP Source | FHIR Source | ...
    """

    ws = wb.create_sheet("Cross-Reference")

    # Discover which sources are actually present in this CDM
    source_types = _discover_source_types(extractor)

    if not source_types:
        # Fallback: use the known list so the tab isn't empty
        source_types = [s for s in _SOURCE_ORDER]

    # Per-ancillary attribute indices — only ancillary sources need this
    # remapping (their entities get renamed by the rationalizer).  EDW /
    # FHIR / NCPDP / Glue / Guardrails already store actual source-side
    # names in source_entity, so they're unchanged.
    ancillary_indices = {}
    if outdir is not None and cdm_name:
        for s in source_types:
            if s.startswith("ancillary"):
                ancillary_indices[s] = ancillary_attribute_index(outdir, cdm_name, s)

    # Build header list
    fixed_headers = ["Entity", "Attribute", "Data Type"]
    source_headers = [_format_source_label(s) for s in source_types]
    headers = fixed_headers + source_headers

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data rows
    row = 2
    attributes = extractor.get_all_attributes()

    for attr in attributes:
        lineage = attr.source_lineage or {}

        source_cells: List[str] = []
        for s in source_types:
            if s.startswith("ancillary"):
                # Use the ancillary attr-index to render real table.column;
                # schema is dropped here because the column header already
                # identifies the ancillary source.
                refs = format_ancillary_source_refs(
                    ancillary_indices.get(s),
                    lineage.get(s, []),
                    include_schema=False,
                )
                source_cells.append("; ".join(refs))
            else:
                source_cells.append(_extract_source_field(lineage, s))

        data = [
            attr.entity_name,
            attr.attribute_name,
            attr.data_type,
        ] + source_cells

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        row += 1

    # Column widths
    fixed_widths = [20, 25, 15]         # Entity, Attribute, Data Type
    source_widths = [35] * len(source_types)
    widths = fixed_widths + source_widths

    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Autofilter across all columns
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{row - 1}"

    return ws