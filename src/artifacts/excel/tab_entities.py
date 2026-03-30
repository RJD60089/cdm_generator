# src/artifacts/excel/tab_entities.py
"""Generate Entities tab for Excel CDM."""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.styles import ExcelStyles


# Preferred display order for source columns.
# Any source not in this list is appended alphabetically.
_SOURCE_ORDER = ["edw", "guardrails", "glue", "ncpdp", "fhir"]
_UPPER_LABELS = {"edw", "ncpdp", "fhir"}


def _source_label(source_type: str) -> str:
    """Convert source_type key to a readable column header."""
    return source_type.upper() if source_type.lower() in _UPPER_LABELS else source_type.title()


def create_entities_tab(wb: Workbook, extractor: CDMExtractor) -> None:
    """
    Create the Entities tab with entity overview.

    Fixed columns:
        Entity Name | Description | Classification | Primary Key(s) | Attribute Count

    Dynamic source columns (one per source type found across entity source_coverage):
        EDW | Guardrails | Glue | NCPDP | FHIR | ...
    """

    ws = wb.create_sheet("Entities")

    entities = extractor.get_entities()

    # --- Discover source types from entity source_coverage ---
    found_sources: set = set()
    for entity in entities:
        found_sources.update(entity.source_coverage.keys())

    # Order: preferred list first, then any extras alphabetically
    source_types = [s for s in _SOURCE_ORDER if s in found_sources]
    source_types += sorted(s for s in found_sources if s not in _SOURCE_ORDER)

    # --- Build headers ---
    fixed_headers = [
        "Entity Name", "Description", "Classification",
        "Primary Key(s)", "Attribute Count"
    ]
    headers = fixed_headers + [_source_label(s) for s in source_types]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        ExcelStyles.apply_header_style(cell)

    # --- Data rows ---
    num_fixed = len(fixed_headers)

    for row_idx, entity in enumerate(entities, 2):
        is_alt = row_idx % 2 == 0

        pk_str = ", ".join(entity.primary_keys) if entity.primary_keys else ""

        row_data = [
            entity.name,
            entity.description,
            entity.classification,
            pk_str,
            entity.attribute_count,
        ] + ["✓" if entity.source_coverage.get(s) else "" for s in source_types]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            ExcelStyles.apply_body_style(cell, is_alt)
            if col > num_fixed:
                cell.alignment = ExcelStyles.CENTER_ALIGN

    # --- Column widths ---
    fixed_widths = [25, 60, 15, 30, 15]
    source_widths = [max(10, len(_source_label(s)) + 2) for s in source_types]

    for i, width in enumerate(fixed_widths + source_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # --- Freeze and filter ---
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(entities) + 1}"