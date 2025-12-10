# src/artifacts/excel/tab_summary.py
"""Generate Summary tab for Excel CDM."""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from src.config.config_parser import AppConfig
from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.gap_extractor import GapExtractor
from src.artifacts.common.styles import ExcelStyles


def create_summary_tab(
    wb: Workbook,
    extractor: CDMExtractor,
    config: AppConfig,
    gap_extractor: GapExtractor
) -> None:
    """
    Create the Summary tab with metadata and counts.
    
    Contents:
    - Domain info, generation date, version
    - Entity/attribute counts
    - Source coverage summary
    - Source priority (survivorship)
    - Gap summary
    """
    
    ws = wb.create_sheet("Summary")
    
    # Title
    ws["A1"] = f"{extractor.domain} CDM Summary"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")
    
    row = 3
    
    # === CDM Information ===
    row = _add_section_header(ws, row, "CDM Information")
    
    info_data = [
        ("Domain", extractor.domain),
        ("Type", config.cdm.type),
        ("Version", extractor.version),
        ("Generated Date", extractor.generated_date[:10] if extractor.generated_date else datetime.now().strftime("%Y-%m-%d")),
        ("Description", extractor.domain_description)
    ]
    
    for label, value in info_data:
        ws.cell(row=row, column=1, value=label).font = ExcelStyles.BOLD_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    row += 1
    
    # === Counts ===
    row = _add_section_header(ws, row, "CDM Statistics")
    
    count_data = [
        ("Total Entities", extractor.entity_count),
        ("Total Attributes", extractor.attribute_count),
        ("Total Relationships", len(extractor.get_relationships()))
    ]
    
    for label, value in count_data:
        ws.cell(row=row, column=1, value=label).font = ExcelStyles.BOLD_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    row += 1
    
    # === Source Coverage ===
    row = _add_section_header(ws, row, "Source Coverage")
    
    coverage = extractor.get_source_coverage_summary()
    
    # Priority order header
    ws.cell(row=row, column=1, value="Source Priority (Survivorship):").font = ExcelStyles.BOLD_FONT
    ws.cell(row=row, column=2, value="Guardrails → Glue → NCPDP → FHIR")
    row += 1
    
    ws.cell(row=row, column=1, value="").font = ExcelStyles.BOLD_FONT  # spacer
    row += 1
    
    # Coverage counts (in priority order)
    for source in ["guardrails", "glue", "ncpdp", "fhir"]:
        count = coverage.get(source, 0)
        label = source.upper() if source in ["ncpdp", "fhir"] else source.title()
        ws.cell(row=row, column=1, value=f"  {label}").font = ExcelStyles.BOLD_FONT
        ws.cell(row=row, column=2, value=f"{count} entities")
        row += 1
    
    row += 1
    
    # === Gap Summary ===
    if gap_extractor.gaps:
        row = _add_section_header(ws, row, "Gap Analysis Summary")
        
        summary = gap_extractor.summary
        gap_data = [
            ("Unmapped Fields", summary.get("total_unmapped", 0)),
            ("Fields Requiring Review", summary.get("total_requires_review", 0))
        ]
        
        for label, value in gap_data:
            ws.cell(row=row, column=1, value=label).font = ExcelStyles.BOLD_FONT
            ws.cell(row=row, column=2, value=value)
            row += 1
    
    row += 1
    
    # === Source Files ===
    row = _add_section_header(ws, row, "Source Files")
    
    source_files = extractor.source_files
    for source_type, filename in source_files.items():
        label = source_type.upper() if source_type in ["ncpdp", "fhir"] else source_type.title()
        ws.cell(row=row, column=1, value=label).font = ExcelStyles.BOLD_FONT
        ws.cell(row=row, column=2, value=filename)
        row += 1
    
    # Column widths
    widths = {
        "A": 30,
        "B": 80,
        "C": 20,
        "D": 20
    }
    ExcelStyles.set_column_widths(ws, widths)


def _add_section_header(ws, row: int, title: str) -> int:
    """Add a section header and return next row."""
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="4472C4")
    ws.merge_cells(f"A{row}:D{row}")
    return row + 1
