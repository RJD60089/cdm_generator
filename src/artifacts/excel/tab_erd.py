# src/artifacts/excel/tab_erd.py
"""Generate ERD tab for Excel CDM."""

from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font
from src.artifacts.common.styles import ExcelStyles


def create_erd_tab(wb: Workbook, erd_url: Optional[str] = None) -> None:
    """
    Create the ERD tab with link to diagram.
    
    Contents:
    - Link to LucidChart or other ERD tool
    - Instructions for viewing
    """
    
    ws = wb.create_sheet("ERD")
    
    # Title
    ws["A1"] = "Entity Relationship Diagram"
    ws["A1"].font = Font(bold=True, size=14)
    
    # Link or placeholder
    ws["A3"] = "ERD Location:"
    ws["A3"].font = ExcelStyles.BOLD_FONT
    
    if erd_url:
        ws["B3"] = erd_url
        ws["B3"].font = ExcelStyles.LINK_FONT
        ws["B3"].hyperlink = erd_url
    else:
        ws["B3"] = "(ERD URL not provided - see LucidChart import in artifacts folder)"
    
    # Instructions
    ws["A5"] = "Instructions:"
    ws["A5"].font = ExcelStyles.BOLD_FONT
    
    instructions = [
        "1. Import the LucidChart CSV file into LucidChart to generate the ERD",
        "2. The CSV file is located in the artifacts folder alongside this Excel file",
        "3. In LucidChart: File → Import Data → Entity Relationship (ERD)",
        "4. Select the lucidchart_*.csv file and follow the import wizard"
    ]
    
    for i, instruction in enumerate(instructions):
        ws.cell(row=6 + i, column=1, value=instruction)
    
    # Notes
    ws["A12"] = "Notes:"
    ws["A12"].font = ExcelStyles.BOLD_FONT
    
    notes = [
        "• The ERD shows all entities and their foreign key relationships",
        "• Primary keys are marked in the import",
        "• Update this cell with the published LucidChart URL after creation"
    ]
    
    for i, note in enumerate(notes):
        ws.cell(row=13 + i, column=1, value=note)
    
    # Column widths
    widths = {"A": 20, "B": 80}
    ExcelStyles.set_column_widths(ws, widths)
