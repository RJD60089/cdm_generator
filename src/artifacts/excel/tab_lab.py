# src/artifacts/excel/tab_lab.py
"""
Generate Lab tabs for CDM Workshop use.

Two worksheets:
  - Data_Dictionary_Lab: Every Data Dictionary column is preserved; the
    workshop columns (Add/Remove/Change/Reference Data, If Conditional,
    Fixed Values, Notes, Comment, Finalized) are inserted immediately
    after "Business Definition" with GREEN headers.
  - Entities_Lab: Every Entities column is preserved; GREEN lab columns
    (Decision, Note, Lab 3 Notes) and ORANGE lab columns (What Is This?,
    Where From?, What Forms?, Where Ends?, Lab 4 Notes) are inserted
    immediately after "Description".
"""

from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.gap_extractor import GapExtractor
from src.artifacts.common.styles import ExcelStyles


# --- Lab header colors ---
GREEN_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
LAB_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LAB_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


# Entities Lab — GREEN columns (2-4 per spec)
ENTITY_GREEN_COLUMNS = [
    ("Decision", "CONFIRM / RENAME / ADD / REMOVE / MOVE / SPLIT-MERGE"),
    ("Note", "(only for ADD, RENAME, MOVE, SPLIT/MERGE)"),
    ("Lab 3 Notes", ""),
]

# Entities Lab — ORANGE columns (5-9 per spec)
ENTITY_ORANGE_COLUMNS = [
    (
        "What Is This?",
        "A plain-language description of this entity, written so any business "
        "person could read it and understand it. One to three sentences. No "
        "technical language, no jargon.",
    ),
    (
        "Where Does It Come From?",
        "What business process creates this entity or causes it to exist? "
        "When does it show up in the business? What triggers its creation?",
    ),
    (
        "What Forms Does It Take?",
        "Are there different types or variations of this entity that behave "
        "differently or mean something different? For example: retail claims, "
        "mail order claims, compound claims, specialty claims.",
    ),
    (
        "Where Does It End?",
        "Where does this entity's responsibility stop and another CDM's "
        "begin? What does this entity NOT include? What belongs elsewhere?",
    ),
    ("Lab 4 Notes", ""),
]


# Data Dictionary Lab — GREEN workshop columns, placed after Business Definition
DD_GREEN_COLUMNS = [
    ("Add/Remove/Change/Reference Data", ""),
    ("If Conditional", ""),
    ("Fixed Values", ""),
    ("Notes", ""),
    ("Comment", ""),
    ("Finalized", ""),
]


# Source type label helpers (mirror tab_entities.py)
_SOURCE_ORDER = ["edw", "guardrails", "glue", "ncpdp", "fhir"]
_UPPER_LABELS = {"edw", "ncpdp", "fhir"}


def _source_label(source_type: str) -> str:
    return source_type.upper() if source_type.lower() in _UPPER_LABELS else source_type.title()


def _apply_lab_header(cell, fill: PatternFill) -> None:
    cell.fill = fill
    cell.font = LAB_HEADER_FONT
    cell.alignment = LAB_HEADER_ALIGN
    cell.border = ExcelStyles.THIN_BORDER


def _compose_header(label: str, note: str) -> str:
    return f"{label}\n{note}" if note else label


# =============================================================================
# DATA DICTIONARY LAB
# =============================================================================

def create_data_dictionary_lab_tab(
    wb: Workbook,
    extractor: CDMExtractor,
    gap_extractor: Optional[GapExtractor] = None,
) -> None:
    """
    Create the Data Dictionary Lab tab — identical to Data Dictionary with
    GREEN workshop columns inserted immediately after Business Definition.

    When ``gap_extractor`` is provided, an ORANGE "Needs Review" column is
    appended immediately after the final GREEN column ("Finalized") and
    is set to "Y" for any CDM attribute that appears on Requires_Review.
    """

    ws = wb.create_sheet("Data_Dictionary_Lab")

    attributes = extractor.get_all_attributes()

    has_field_codes = any(a.ncpdp_field_codes or a.edw_field_codes for a in attributes)

    ancillary_sources = sorted({
        key for a in attributes
        for key in a.source_lineage
        if key.startswith("ancillary") and a.source_lineage[key]
    })

    # Build the set of (entity, attribute) pairs that have any SME-review
    # mappings flagged in the gap analysis.
    review_pairs: set = set()
    if gap_extractor is not None:
        for f in gap_extractor.get_requires_review_fields():
            if f.cdm_entity and f.cdm_attribute:
                review_pairs.add((f.cdm_entity, f.cdm_attribute))

    # Headers — Entity, Attribute, Business Definition, GREEN lab cols,
    # then an optional ORANGE "Needs Review" flag, then remaining DD cols.
    tail_headers = [
        "Data Type", "Size", "Nullable", "Is PK", "Is FK", "FK Reference",
        "Classification", "PII", "PHI", "Rematch",
    ]
    if has_field_codes:
        tail_headers += ["NCPDP Field Code", "EDW F-Code"]
    for anc_src in ancillary_sources:
        display_name = anc_src.replace("ancillary-", "").replace("-", " ").title()
        tail_headers.append(f"Ancillary {display_name}")

    header_specs = (
        [("Entity", None), ("Attribute", None), ("Business Definition", None)]
        + [(_compose_header(l, n), "green") for l, n in DD_GREEN_COLUMNS]
    )
    if gap_extractor is not None:
        header_specs.append(("Needs Review", "orange"))
    header_specs += [(h, None) for h in tail_headers]

    for col_idx, (header, color) in enumerate(header_specs, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if color == "green":
            _apply_lab_header(cell, GREEN_FILL)
        elif color == "orange":
            _apply_lab_header(cell, ORANGE_FILL)
        else:
            ExcelStyles.apply_header_style(cell)

    num_green = len(DD_GREEN_COLUMNS)
    has_review_col = gap_extractor is not None

    for row_idx, attr in enumerate(attributes, 2):
        is_alt = row_idx % 2 == 0

        size = ""
        if attr.max_length:
            size = str(attr.max_length)
        elif attr.precision:
            size = f"{attr.precision},{attr.scale or 0}"

        is_rematch = any(
            mapping.get("rematch") is True
            for mappings in attr.source_lineage.values()
            for mapping in (mappings if isinstance(mappings, list) else [])
        )

        tail_data = [
            attr.data_type,
            size,
            "Y" if attr.nullable else "N",
            "Y" if attr.pk else "",
            "Y" if attr.fk_to else "",
            attr.fk_to or "",
            attr.classification or "",
            "Y" if attr.is_pii else "",
            "Y" if attr.is_phi else "",
            "R" if is_rematch else "",
        ]
        if has_field_codes:
            tail_data += [
                "; ".join(attr.ncpdp_field_codes) if attr.ncpdp_field_codes else "",
                "; ".join(attr.edw_field_codes) if attr.edw_field_codes else "",
            ]
        for anc_src in ancillary_sources:
            entries = attr.source_lineage.get(anc_src, [])
            refs = []
            if isinstance(entries, list):
                for e in entries:
                    src_entity = e.get("source_entity", "")
                    src_attr = e.get("source_attribute", "")
                    if src_entity and src_attr:
                        refs.append(f"{src_entity}.{src_attr}")
                    elif src_attr:
                        refs.append(src_attr)
            tail_data.append("; ".join(refs))

        needs_review_val = []
        if has_review_col:
            needs_review_val = [
                "Y" if (attr.entity_name, attr.attribute_name) in review_pairs else ""
            ]

        row_data = (
            [attr.entity_name, attr.attribute_name, attr.description or ""]
            + [""] * num_green
            + needs_review_val
            + tail_data
        )

        # Column index where the Rematch cell lives in this tab. "Rematch"
        # is the 10th entry in tail_data, and tail_data starts right after
        # the Needs-Review column (when present).
        rematch_col = 3 + num_green + (1 if has_review_col else 0) + 10

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            ExcelStyles.apply_body_style(cell, is_alt)
            if col_idx == 2:
                if attr.pk:
                    ExcelStyles.apply_pk_style(cell)
                elif attr.fk_to:
                    ExcelStyles.apply_fk_style(cell)
            if col_idx == rematch_col and is_rematch:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                cell.font = Font(bold=True, color="7D6608")

    # Column widths
    widths = {"A": 25, "B": 30, "C": 50}
    for i in range(num_green):
        widths[get_column_letter(4 + i)] = 22
    tail_start = 4 + num_green
    if has_review_col:
        widths[get_column_letter(tail_start)] = 14   # "Needs Review"
        tail_start += 1

    tail_widths_fixed = [15, 10, 10, 8, 8, 35, 15, 8, 8, 10]
    for i, w in enumerate(tail_widths_fixed):
        widths[get_column_letter(tail_start + i)] = w
    next_col = tail_start + len(tail_widths_fixed)
    if has_field_codes:
        widths[get_column_letter(next_col)] = 20
        widths[get_column_letter(next_col + 1)] = 20
        next_col += 2
    for i in range(len(ancillary_sources)):
        widths[get_column_letter(next_col + i)] = 30
    ExcelStyles.set_column_widths(ws, widths)

    ws.row_dimensions[1].height = 40

    ws.freeze_panes = "D2"
    last_col = get_column_letter(len(header_specs))
    ws.auto_filter.ref = f"A1:{last_col}{len(attributes) + 1}"


# =============================================================================
# ENTITIES LAB
# =============================================================================

def create_entities_lab_tab(wb: Workbook, extractor: CDMExtractor) -> None:
    """
    Create the Entities Lab tab — identical to Entities with GREEN and
    ORANGE lab columns inserted immediately after Description. The
    remaining Entities columns (Classification, Primary Key(s),
    Attribute Count, source coverage) follow the lab columns.
    """

    ws = wb.create_sheet("Entities_Lab")

    entities = extractor.get_entities()

    found_sources = set()
    for entity in entities:
        found_sources.update(entity.source_coverage.keys())
    source_types = [s for s in _SOURCE_ORDER if s in found_sources]
    source_types += sorted(s for s in found_sources if s not in _SOURCE_ORDER)

    # Headers: Entity Name, Description, GREEN lab cols, ORANGE lab cols,
    # then the rest of the Entities columns.
    tail_headers = ["Classification", "Primary Key(s)", "Attribute Count"] + [
        _source_label(s) for s in source_types
    ]

    header_specs = (
        [("Entity Name", None), ("Description", None)]
        + [(_compose_header(l, n), "green") for l, n in ENTITY_GREEN_COLUMNS]
        + [(_compose_header(l, n), "orange") for l, n in ENTITY_ORANGE_COLUMNS]
        + [(h, None) for h in tail_headers]
    )

    for col_idx, (header, color) in enumerate(header_specs, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if color == "green":
            _apply_lab_header(cell, GREEN_FILL)
        elif color == "orange":
            _apply_lab_header(cell, ORANGE_FILL)
        else:
            ExcelStyles.apply_header_style(cell)

    num_green = len(ENTITY_GREEN_COLUMNS)
    num_orange = len(ENTITY_ORANGE_COLUMNS)
    num_lab = num_green + num_orange
    num_fixed_tail = 3  # Classification, Primary Key(s), Attribute Count

    for row_idx, entity in enumerate(entities, 2):
        is_alt = row_idx % 2 == 0

        pk_str = ", ".join(entity.primary_keys) if entity.primary_keys else ""

        row_data = (
            [entity.name, entity.description]
            + [""] * num_lab
            + [
                entity.classification,
                pk_str,
                entity.attribute_count,
            ]
            + ["✓" if entity.source_coverage.get(s) else "" for s in source_types]
        )

        source_cols_start = 2 + num_lab + num_fixed_tail + 1

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            ExcelStyles.apply_body_style(cell, is_alt)
            if col_idx >= source_cols_start:
                cell.alignment = ExcelStyles.CENTER_ALIGN

    # Column widths
    widths = {"A": 25, "B": 60}
    for i in range(num_green):
        widths[get_column_letter(3 + i)] = 22
    for i in range(num_orange):
        widths[get_column_letter(3 + num_green + i)] = 35
    tail_start = 3 + num_lab
    fixed_widths = [15, 30, 15]
    for i, w in enumerate(fixed_widths):
        widths[get_column_letter(tail_start + i)] = w
    src_start = tail_start + num_fixed_tail
    for i, s in enumerate(source_types):
        widths[get_column_letter(src_start + i)] = max(10, len(_source_label(s)) + 2)
    ExcelStyles.set_column_widths(ws, widths)

    ws.row_dimensions[1].height = 90

    ws.freeze_panes = "C2"
    last_col = get_column_letter(len(header_specs))
    ws.auto_filter.ref = f"A1:{last_col}{len(entities) + 1}"
