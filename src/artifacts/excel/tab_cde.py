# src/artifacts/excel/tab_cde.py
"""Generate Candidate CDEs tab for Excel CDM."""

from openpyxl import Workbook
from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.styles import ExcelStyles


def create_cde_tab(wb: Workbook, extractor: CDMExtractor) -> None:
    """
    Create the Candidate CDEs tab.
    
    Reads CDEs directly from cdm["critical_data_elements"] 
    (populated by postprocess_cde.py in Step 6).
    
    Columns:
    - Entity, Attribute, CDE Category, Business Justification
    - Data Type, PII, PHI
    """
    
    ws = wb.create_sheet("Candidate_CDEs")
    
    # Headers
    headers = [
        "Entity", "Attribute", "CDE Category", 
        "Business Justification", "Data Type", "PII", "PHI"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        ExcelStyles.apply_header_style(cell)
    
    # Get CDEs directly from CDM
    cdes = extractor.cdm.get("critical_data_elements", [])
    
    # Build attribute lookup for metadata (data_type, is_pii, is_phi)
    attr_lookup = {}
    for entity in extractor.cdm.get("entities", []):
        entity_name = entity.get("entity_name", "")
        for attr in entity.get("attributes", []):
            attr_name = attr.get("attribute_name", "")
            attr_lookup[(entity_name, attr_name)] = attr
    
    # Data rows
    for row_idx, cde in enumerate(cdes, 2):
        is_alt = row_idx % 2 == 0
        
        entity = cde.get("entity", "")
        attribute = cde.get("attribute", "")
        category = cde.get("cde_category", "")
        justification = cde.get("justification", "")
        
        # Get attribute metadata
        attr_meta = attr_lookup.get((entity, attribute), {})
        data_type = attr_meta.get("data_type", "")
        is_pii = attr_meta.get("is_pii", False)
        is_phi = attr_meta.get("is_phi", False)
        
        row_data = [
            entity,
            attribute,
            category,
            justification,
            data_type,
            "Y" if is_pii else "",
            "Y" if is_phi else ""
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            ExcelStyles.apply_body_style(cell, is_alt)
    
    # Column widths
    widths = {
        "A": 25,  # Entity
        "B": 30,  # Attribute
        "C": 25,  # CDE Category
        "D": 60,  # Business Justification
        "E": 15,  # Data Type
        "F": 8,   # PII
        "G": 8    # PHI
    }
    ExcelStyles.set_column_widths(ws, widths)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Auto-filter
    if cdes:
        ws.auto_filter.ref = f"A1:G{len(cdes) + 1}"
