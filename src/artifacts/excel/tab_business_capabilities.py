# src/artifacts/excel/tab_business_capabilities.py
"""Generate Business Capabilities tab for Excel CDM."""

from openpyxl import Workbook
from src.config.config_parser import AppConfig
from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.styles import ExcelStyles


def create_business_capabilities_tab(
    wb: Workbook, 
    extractor: CDMExtractor,
    config: AppConfig
) -> None:
    """
    Create the Business Capabilities tab.
    
    Derives capabilities from entity classifications and config.
    
    Columns:
    - Business Capability, Description, CDM Tables Involved, Business Value
    """
    
    ws = wb.create_sheet("Business_Capabilities")
    
    # Headers
    headers = [
        "Business Capability", "Description", 
        "CDM Tables Involved", "Business Value"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        ExcelStyles.apply_header_style(cell)
    
    # Derive capabilities from entities and domain
    capabilities = _derive_capabilities(extractor, config)
    
    # Data rows
    for row_idx, cap in enumerate(capabilities, 2):
        is_alt = row_idx % 2 == 0
        
        row_data = [
            cap["capability"],
            cap["description"],
            cap["tables"],
            cap["value"]
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            ExcelStyles.apply_body_style(cell, is_alt)
    
    # Column widths
    widths = {
        "A": 30,  # Capability
        "B": 60,  # Description
        "C": 40,  # Tables
        "D": 50   # Business Value
    }
    ExcelStyles.set_column_widths(ws, widths)
    
    # Freeze header row
    ws.freeze_panes = "A2"


def _derive_capabilities(extractor: CDMExtractor, config: AppConfig) -> list:
    """Derive business capabilities from CDM structure."""
    
    capabilities = []
    entities = extractor.get_entities()
    domain = extractor.domain
    
    # Group entities by classification
    core_entities = [e.name for e in entities if e.classification == "Core"]
    ref_entities = [e.name for e in entities if e.classification == "Reference"]
    trans_entities = [e.name for e in entities if e.classification == "Transactional"]
    
    # Core capability - always present
    capabilities.append({
        "capability": f"{domain} Identity Management",
        "description": f"Manage core {domain.lower()} definitions, identifiers, and lifecycle",
        "tables": ", ".join(core_entities[:5]) + ("..." if len(core_entities) > 5 else ""),
        "value": f"Single source of truth for {domain.lower()} master data"
    })
    
    # Hierarchy capability if multiple related entities
    if len(entities) > 3:
        capabilities.append({
            "capability": f"{domain} Hierarchy Management",
            "description": f"Maintain organizational and structural hierarchies within {domain.lower()} domain",
            "tables": ", ".join([e.name for e in entities if "group" in e.name.lower() or 
                                "organization" in e.name.lower() or "association" in e.name.lower()][:5]),
            "value": "Support complex organizational structures and relationships"
        })
    
    # Relationship tracking
    relationships = extractor.get_relationships()
    if relationships:
        capabilities.append({
            "capability": "Relationship Tracking",
            "description": "Track relationships between entities across the domain",
            "tables": ", ".join(set([r.child_entity for r in relationships[:5]])),
            "value": "Enable cross-entity analytics and lineage"
        })
    
    # Source integration
    coverage = extractor.get_source_coverage_summary()
    sources = [s for s, c in coverage.items() if c > 0]
    if sources:
        capabilities.append({
            "capability": "Multi-Source Integration",
            "description": f"Integrate data from {', '.join(sources)}",
            "tables": "All entities",
            "value": "Unified view across internal and external data sources"
        })
    
    # Compliance/Governance
    capabilities.append({
        "capability": "Data Governance",
        "description": "Support data quality, lineage tracking, and compliance requirements",
        "tables": "All entities (audit columns)",
        "value": "Regulatory compliance and audit readiness"
    })
    
    return capabilities
