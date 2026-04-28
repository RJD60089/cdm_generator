# src/artifacts/excel/generate_excel_cdm.py
"""
Excel CDM Generator - Main Orchestrator

Generates a complete Excel CDM workbook with 13 tabs.
"""

from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional
from openpyxl import Workbook

from src.config.config_parser import AppConfig
from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.gap_extractor import GapExtractor

# Tab generators
from src.artifacts.excel.tab_data_dictionary import create_data_dictionary_tab
from src.artifacts.excel.tab_entities import create_entities_tab
from src.artifacts.excel.tab_relationships import create_relationships_tab
from src.artifacts.excel.tab_cross_reference import create_cross_reference_tab
from src.artifacts.excel.tab_cde import create_cde_tab
from src.artifacts.excel.tab_business_rules import create_business_rules_tab
from src.artifacts.excel.tab_business_rules_consolidated import create_business_rules_consolidated_tab
from src.artifacts.excel.tab_business_capabilities import create_business_capabilities_tab
from src.artifacts.excel.tab_summary import create_summary_tab
from src.artifacts.excel.tab_erd import create_erd_tab
from src.artifacts.excel.tab_requires_review import create_requires_review_tab
from src.artifacts.excel.tab_sme_questions import create_sme_questions_tab
from src.artifacts.excel.tab_unmapped import create_unmapped_tab
from src.artifacts.excel.tab_source_files import create_source_files_tab
from src.artifacts.excel.tab_lab import (
    create_data_dictionary_lab_tab,
    create_entities_lab_tab,
)
from src.artifacts.excel.tab_mapping import create_mapping_tab


def generate_excel_cdm(
    config: AppConfig,
    cdm_path: Path,
    output_path: Path,
    gaps_path: Optional[Path] = None,
    consolidation_path: Optional[Path] = None,
    erd_url: Optional[str] = None,
    consolidated_rules_path: Optional[Path] = None,
) -> Path:
    """
    Generate complete Excel CDM workbook.
    
    Args:
        config: Application configuration
        cdm_path: Path to Full CDM JSON
        output_path: Output Excel file path
        gaps_path: Optional path to gaps JSON
        consolidation_path: Optional path to consolidation recommendations JSON
        erd_url: Optional URL to ERD diagram
    
    Returns:
        Path to generated Excel file
    """
    
    print(f"   Loading CDM data...")
    
    # Initialize extractors
    extractor = CDMExtractor(cdm_path=cdm_path)
    gap_extractor = None
    if gaps_path and gaps_path.exists():
        gap_extractor = GapExtractor(
            gaps_path=gaps_path,
            consolidation_path=consolidation_path
        )
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Tab order (13 tabs total)
    print(f"   Creating tabs...")
    
    # 1. Summary - overview first
    print(f"      - Summary")
    create_summary_tab(wb, extractor, config, gap_extractor)
    
    # 2. Entities - entity overview
    print(f"      - Entities")
    create_entities_tab(wb, extractor)

    # 2b. Entities Lab - workshop columns for entity-level review
    print(f"      - Entities_Lab")
    create_entities_lab_tab(wb, extractor)

    # 3. Data Dictionary - primary reference
    print(f"      - Data_Dictionary")
    create_data_dictionary_tab(wb, extractor)

    # 3b. Data Dictionary Lab - workshop columns for attribute-level review
    print(f"      - Data_Dictionary_Lab")
    create_data_dictionary_lab_tab(wb, extractor, gap_extractor)

    # 3c. Mapping — Collibra source-to-target mapping, placed
    # immediately after Data_Dictionary_Lab so attribute-level tabs
    # stay grouped.
    # Schema-resolver lookups read from outdir/rationalized/ and from the
    # ancillary source files referenced by config — so pass the base outdir.
    print(f"      - Mapping")
    create_mapping_tab(wb, extractor, config, outdir=cdm_path.parent.parent)

    # 3d. Candidate CDEs - placed after Mapping so the CDE list sits
    # alongside the attribute-level tabs and can be referenced by the
    # Mapping tab.
    print(f"      - Candidate_CDEs")
    create_cde_tab(wb, extractor)

    # 4. Relationships - FK details
    print(f"      - Relationships")
    create_relationships_tab(wb, extractor)

    # 5. Cross-Reference - source lineage mapping
    print(f"      - Cross_Reference")
    create_cross_reference_tab(wb, extractor)

    # 7. Business Rules
    print(f"      - Business_Rules")
    create_business_rules_tab(wb, extractor)

    # 7b. Business Rules Consolidated — AI-driven (reads separate JSON)
    print(f"      - Business_Rules_Consolidated")
    create_business_rules_consolidated_tab(wb, consolidated_rules_path)

    # 8. Business Capabilities
    print(f"      - Business_Capabilities")
    create_business_capabilities_tab(wb, extractor, config)
    
    # 9. Requires Review - from gap analysis
    print(f"      - Requires_Review")
    create_requires_review_tab(wb, extractor, gap_extractor)
    
    # 10. SME Questions - from gap analysis
    print(f"      - SME_Questions")
    create_sme_questions_tab(wb, extractor, gap_extractor)
    
    # 11. Unmapped Fields - from gap analysis
    print(f"      - Unmapped_Fields")
    create_unmapped_tab(wb, extractor, gap_extractor)
    
    # 12. Source Files - config inputs
    print(f"      - Source_Files")
    create_source_files_tab(wb, config)
    
    # 13. ERD
    print(f"      - ERD")
    create_erd_tab(wb, erd_url)

    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    print(f"   Excel CDM saved: {output_path.name}")
    print(f"      Entities: {extractor.entity_count}")
    print(f"      Attributes: {extractor.attribute_count}")
    print(f"      Tabs: {len(wb.sheetnames)}")
    
    return output_path


def find_gaps_file(outdir: Path, domain: str) -> Optional[Path]:
    """Find latest gaps file."""
    domain_safe = domain.lower().replace(' ', '_')
    full_cdm_dir = outdir / "full_cdm"
    
    if not full_cdm_dir.exists():
        return None
    
    pattern = f"gaps_{domain_safe}_*.json"
    matches = list(full_cdm_dir.glob(pattern))
    
    if not matches:
        return None
    
    matches.sort(reverse=True)
    return matches[0]


def find_consolidation_file(outdir: Path, domain: str) -> Optional[Path]:
    """Find latest consolidation recommendations file."""
    domain_safe = domain.lower().replace(' ', '_')
    cdm_dir = outdir / "cdm"
    
    if not cdm_dir.exists():
        return None
    
    pattern = f"consolidation_recommendations_{domain_safe}_*.json"
    matches = list(cdm_dir.glob(pattern))
    
    if not matches:
        return None
    
    matches.sort(reverse=True)
    return matches[0]
