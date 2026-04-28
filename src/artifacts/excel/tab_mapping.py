# src/artifacts/excel/tab_mapping.py
"""
Generate the Mapping tab for Collibra loads.

Every row is a single source-to-target mapping between one source
<table>.<column> and one CDM entity.attribute.  Sources are drawn from
the lineage keys listed in config.mapping.mapping_sources (e.g. "edw",
"ancillary-cost-share-accumulators").

Row rules (from requirements):
  - Clear duplicate — same source table.column + same target
    entity.attribute across two mapping sources → single merged row with
    both source Y/N columns = Y.
  - Different source matches for the same CDM attribute → separate rows.
  - CDM attribute with no mappings from any configured source → one row
    with blank source fields and all Y/N = N.

Styling: alternating highlight by target entity.attribute group.  All
rows for the same CDM attribute share a shade; the shade flips at each
new attribute.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.schema_resolver import SchemaResolver, ancillary_attribute_index
from src.artifacts.common.styles import ExcelStyles
from src.config.config_parser import AppConfig


# Group highlight colors — two shades that alternate per CDM attribute
GROUP_FILL_A = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GROUP_FILL_B = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")


FIXED_HEADERS = [
    "Source Application",
    "Source Schema",
    "Source Table",
    "Source Column",
    "Source Full Name",
    "Target CDM",
    "Target Table",
    "Target Column",
    "Target Full Name",
    "Description",
    "SQL Type",
    "Iceberg Type",
    "Python Type",
    "PII",
    "PHI",
    "CDE",
]


# =============================================================================
# TYPE DERIVATION
# =============================================================================

_SQL_TYPE_MAP = [
    # (regex, iceberg template, python template).
    # Template may reference match groups as {0}, {1}, ... (the exact groups()
    # tuple is passed to .format()).  Order matters — more specific patterns
    # first so "datetime" wins over "date", "timestamp" wins over "time".
    (r"^\s*(varchar|char|nvarchar|nchar|text|clob|string)", "string", "str"),
    (r"^\s*(bigint|long|int8)", "long", "int"),
    (r"^\s*(smallint|tinyint|int2|int4)", "int", "int"),
    (r"^\s*(int|integer)", "int", "int"),
    (r"^\s*(?:decimal|numeric|number)\s*\(\s*(\d+)\s*,\s*(\d+)\s*\)", "decimal({0},{1})", "Decimal"),
    (r"^\s*(?:decimal|numeric|number)\s*\(\s*(\d+)\s*\)", "decimal({0},0)", "Decimal"),
    (r"^\s*(?:decimal|numeric|number)\b", "decimal(38,18)", "Decimal"),
    (r"^\s*float",   "float",  "float"),
    (r"^\s*real",    "float",  "float"),
    (r"^\s*double",  "double", "float"),
    (r"^\s*(boolean|bit|bool)\b", "boolean", "bool"),
    (r"^\s*datetime\b",  "timestamp", "datetime"),
    (r"^\s*timestamp\b", "timestamp", "datetime"),
    (r"^\s*date\b",      "date",      "date"),
    (r"^\s*time\b",      "time",      "time"),
    (r"^\s*(binary|varbinary|blob|bytea)\b", "binary", "bytes"),
    (r"^\s*uuid\b", "string", "str"),
    (r"^\s*json\b", "string", "str"),
]


def _derive_types(sql_type: str) -> Tuple[str, str]:
    """Return (iceberg_type, python_type) from a CDM SQL type string."""
    if not sql_type:
        return "", ""
    low = sql_type.lower().strip()
    for pattern, iceberg_tpl, python_tpl in _SQL_TYPE_MAP:
        m = re.match(pattern, low)
        if m:
            try:
                iceberg = iceberg_tpl.format(*m.groups())
            except (IndexError, KeyError):
                iceberg = iceberg_tpl
            return iceberg, python_tpl
    return sql_type, "str"


# =============================================================================
# ROW ASSEMBLY
# =============================================================================

def _lineage_entries(
    attr,
    source_key: str,
    ancillary_index: Optional[Dict[Tuple[str, str], List[Dict[str, str]]]] = None,
) -> List[Dict[str, Any]]:
    """
    Normalised list of source-mapping entries for a CDM attribute under
    one source key.  Each returned dict carries:
        - table  : source table name to render in the Mapping row
        - column : source column name to render
        - schema : resolved schema (empty string defers to SchemaResolver)
        - rationalized_entity / rationalized_attribute : original
          source_lineage entry values, kept so callers can still report
          which lineage row produced the mapping if needed

    For ancillary sources, an attribute index (from rationalized JSON's
    per-attribute ``source_attribute`` list) is consulted to recover the
    original ``schema.table.column`` references — the rationalizer
    renames entities to business-friendly names but keeps original
    refs stashed elsewhere; this routine reads those.

    For all other sources (EDW, FHIR, NCPDP, ...), behaviour is
    unchanged: the lineage's ``source_entity`` / ``source_attribute``
    are used directly and the schema is left empty for SchemaResolver
    to fill via the per-source schema lookup.
    """
    entries = attr.source_lineage.get(source_key, []) or []
    if isinstance(entries, dict):
        entries = [entries]

    use_index = (
        ancillary_index is not None
        and source_key.startswith("ancillary")
    )

    out: List[Dict[str, Any]] = []
    for e in entries:
        if not isinstance(e, dict):
            continue
        rationalized_ent = e.get("source_entity", "") or ""
        rationalized_attr = e.get("source_attribute", "") or ""

        if use_index:
            originals = ancillary_index.get(
                (rationalized_ent.lower(), rationalized_attr.lower()),
                [],
            )
            if originals:
                for o in originals:
                    out.append({
                        "table":  o.get("table", "")  or rationalized_ent,
                        "column": o.get("column", "") or rationalized_attr,
                        "schema": o.get("schema", ""),
                        "rationalized_entity":    rationalized_ent,
                        "rationalized_attribute": rationalized_attr,
                    })
                continue

        # Default / fallback: use rationalized values directly, no schema
        if rationalized_ent or rationalized_attr:
            out.append({
                "table":  rationalized_ent,
                "column": rationalized_attr,
                "schema": "",
                "rationalized_entity":    rationalized_ent,
                "rationalized_attribute": rationalized_attr,
            })
    return out


def _build_rows_for_attribute(
    attr,
    mapping_sources: List[str],
    cde_lookup: set,
    cdm_name: str,
    source_application: str,
    schema_resolver: SchemaResolver,
    ancillary_indices: Optional[Dict[str, Dict[Tuple[str, str], List[Dict[str, str]]]]] = None,
) -> List[Dict[str, Any]]:
    """
    Produce one or more row dicts for a single CDM attribute, honouring the
    merge / split rules.

    Returns rows with all 16 fixed columns plus one Y/N per mapping source
    keyed as ``map:<source_key>``.
    """
    target_table = attr.entity_name
    target_col = attr.attribute_name
    target_full = f"{cdm_name}.{target_table}.{target_col}" if cdm_name else f"{target_table}.{target_col}"

    sql_type = attr.data_type or ""
    if attr.max_length:
        sql_type_display = f"{sql_type}({attr.max_length})" if sql_type and "(" not in sql_type else sql_type
    elif attr.precision:
        sql_type_display = (
            f"{sql_type}({attr.precision},{attr.scale or 0})"
            if sql_type and "(" not in sql_type else sql_type
        )
    else:
        sql_type_display = sql_type
    iceberg_type, python_type = _derive_types(sql_type_display)

    description = attr.description or ""
    pii = "Y" if attr.is_pii else "N"
    phi = "Y" if attr.is_phi else "N"
    cde = "Y" if (attr.entity_name, attr.attribute_name) in cde_lookup else "N"

    # key = (source_table, source_column) → {"sources": set of source_keys, "schema": str}
    merged: Dict[Tuple[str, str], Dict[str, Any]] = {}
    indices = ancillary_indices or {}
    for src_key in mapping_sources:
        idx = indices.get(src_key)
        for entry in _lineage_entries(attr, src_key, ancillary_index=idx):
            key = (entry["table"], entry["column"])
            bucket = merged.setdefault(key, {
                "table":  entry["table"],
                "column": entry["column"],
                "schema": entry.get("schema", "") or "",
                "sources": set(),
            })
            bucket["sources"].add(src_key)
            # First non-empty schema wins; subsequent identical hits are fine
            if not bucket["schema"] and entry.get("schema"):
                bucket["schema"] = entry["schema"]

    def _blank_source_flags():
        return {f"map:{s}": "N" for s in mapping_sources}

    rows: List[Dict[str, Any]] = []

    def _row(src_table: str, src_column: str, contributing: Iterable[str], schema_hint: str = "") -> Dict[str, Any]:
        contrib = list(contributing)
        # Schema precedence: explicit per-row hint (from ancillary index)
        # → SchemaResolver.resolve per source → resolver fallback.
        resolved_schema = (schema_hint or "").strip()
        if not resolved_schema and src_table:
            for src_key in contrib:
                resolved_schema = schema_resolver.resolve(src_key, src_table)
                if resolved_schema:
                    break
            if not resolved_schema:
                resolved_schema = schema_resolver.resolve("", src_table)
        src_full = ""
        if src_table or src_column:
            parts = [p for p in [resolved_schema, src_table, src_column] if p]
            src_full = ".".join(parts)
        row = {
            "Source Application": source_application if (src_table or src_column) else "",
            "Source Schema": resolved_schema if (src_table or src_column) else "",
            "Source Table": src_table,
            "Source Column": src_column,
            "Source Full Name": src_full,
            "Target CDM": cdm_name,
            "Target Table": target_table,
            "Target Column": target_col,
            "Target Full Name": target_full,
            "Description": description,
            "SQL Type": sql_type_display,
            "Iceberg Type": iceberg_type,
            "Python Type": python_type,
            "PII": pii,
            "PHI": phi,
            "CDE": cde,
            # group key used for alternating highlights
            "_group_key": f"{target_table}.{target_col}",
        }
        flags = _blank_source_flags()
        for src in contributing:
            flags[f"map:{src}"] = "Y"
        row.update(flags)
        return row

    if not merged:
        rows.append(_row("", "", []))
    else:
        for (tbl, col), bucket in merged.items():
            rows.append(_row(
                tbl, col, bucket["sources"],
                schema_hint=bucket.get("schema", ""),
            ))

    return rows


# =============================================================================
# TAB ENTRY POINT
# =============================================================================

def create_mapping_tab(
    wb: Workbook,
    extractor: CDMExtractor,
    config: AppConfig,
    outdir: Optional["Path"] = None,
) -> None:
    """
    Create the Mapping tab for Collibra ingestion.

    Reads `config.mapping` for source_application, source_schema (used as
    fallback only), and the ordered list of mapping_sources (lineage
    keys).  When ``outdir`` is provided, the SchemaResolver auto-extracts
    per-row schemas:
      - "edw" rows → from rationalized_edw_<domain>_*.json
      - "ancillary-*" DDL rows → parsed from the source DDL file
      - everything else → falls back to ``mapping.source_schema``

    If mapping_sources is empty, emits a one-row placeholder notice.
    """
    ws = wb.create_sheet("Mapping")

    mapping_cfg = config.mapping
    source_application = mapping_cfg.source_application or ""
    mapping_sources = list(mapping_cfg.mapping_sources or [])
    cdm_name = config.cdm.domain or ""

    # Auto-extracting schema resolver. Only useful when we know outdir
    # (where the rationalized JSON lives).  When outdir is None, resolver
    # always returns the config fallback.
    if outdir is None:
        outdir = Path(".")
    schema_resolver = SchemaResolver(config, outdir)

    # Build CDE lookup from cdm["critical_data_elements"]
    cde_list = extractor.cdm.get("critical_data_elements", []) or []
    cde_lookup = {
        (c.get("entity", ""), c.get("attribute", ""))
        for c in cde_list
        if c.get("entity") and c.get("attribute")
    }

    # Headers: fixed columns + one "Mapped: <source>" Y/N per mapping_source
    source_headers = [f"Mapped: {s}" for s in mapping_sources]
    headers = FIXED_HEADERS + source_headers

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        ExcelStyles.apply_header_style(cell)

    # If no mapping sources configured, show placeholder and return
    if not mapping_sources:
        cell = ws.cell(
            row=2, column=1,
            value=(
                "No mapping sources configured. Add 'mapping.mapping_sources' "
                "to config.json (e.g. [\"edw\", \"ancillary-<name>\"]) along "
                "with 'source_application' and 'source_schema', then "
                "regenerate this workbook."
            ),
        )
        cell.font = Font(italic=True, color="7F7F7F", size=10)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        _apply_widths(ws, mapping_sources)
        ws.freeze_panes = "A2"
        return

    # Build per-ancillary attribute indices: maps (rationalized_entity,
    # rationalized_attribute) -> list of {schema, table, column} tuples
    # parsed from the rationalized JSON's source_attribute list.  Used
    # below to expand each ancillary lineage entry into rows that show
    # the original source schema/table/column instead of the rationalized
    # business-friendly entity name.
    ancillary_indices: Dict[str, Dict[Tuple[str, str], List[Dict[str, str]]]] = {}
    for src_key in mapping_sources:
        if src_key.startswith("ancillary"):
            idx = ancillary_attribute_index(outdir, cdm_name, src_key)
            ancillary_indices[src_key] = idx
            if idx:
                print(f"      Ancillary attr-index: {src_key} -> {len(idx)} (entity, attr) pairs")

    # Generate rows, sorted by (entity, attribute) so groups stay contiguous
    attributes = sorted(
        extractor.get_all_attributes(),
        key=lambda a: (a.entity_name, a.attribute_name),
    )

    all_rows: List[Dict[str, Any]] = []
    for attr in attributes:
        all_rows.extend(_build_rows_for_attribute(
            attr,
            mapping_sources=mapping_sources,
            cde_lookup=cde_lookup,
            cdm_name=cdm_name,
            source_application=source_application,
            schema_resolver=schema_resolver,
            ancillary_indices=ancillary_indices,
        ))

    # Diagnostic: report which sources contributed schemas
    stats = schema_resolver.stats()
    if stats:
        for src, n in stats.items():
            print(f"      Schema lookup: {src} -> {n} entities")

    # Sort within-group by source table/column so consistent ordering
    all_rows.sort(key=lambda r: (
        r["_group_key"],
        r["Source Table"] or "~",
        r["Source Column"] or "~",
    ))

    # Write rows with group-based alternating highlight
    prev_group = None
    use_a = True
    for row_idx, row in enumerate(all_rows, 2):
        group = row["_group_key"]
        if group != prev_group:
            use_a = not use_a
            prev_group = group
        fill = GROUP_FILL_A if use_a else GROUP_FILL_B

        for col_idx, header in enumerate(headers, 1):
            value = row.get(header)
            if value is None:
                # Y/N flag lookup via "map:<source>" key
                if header.startswith("Mapped: "):
                    src = header[len("Mapped: "):]
                    value = row.get(f"map:{src}", "N")
                else:
                    value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = ExcelStyles.BODY_FONT
            cell.alignment = ExcelStyles.BODY_ALIGN
            cell.border = ExcelStyles.THIN_BORDER
            cell.fill = fill

    _apply_widths(ws, mapping_sources)
    ws.freeze_panes = "A2"
    if all_rows:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(all_rows) + 1}"


def _apply_widths(ws, mapping_sources: List[str]) -> None:
    widths = {
        "A": 20,  # Source Application
        "B": 18,  # Source Schema
        "C": 25,  # Source Table
        "D": 25,  # Source Column
        "E": 45,  # Source Full Name
        "F": 15,  # Target CDM
        "G": 25,  # Target Table
        "H": 25,  # Target Column
        "I": 45,  # Target Full Name
        "J": 50,  # Description
        "K": 18,  # SQL Type
        "L": 18,  # Iceberg Type
        "M": 14,  # Python Type
        "N": 6,   # PII
        "O": 6,   # PHI
        "P": 6,   # CDE
    }
    ExcelStyles.set_column_widths(ws, widths)
    # Y/N columns
    start = len(FIXED_HEADERS) + 1
    for i, src in enumerate(mapping_sources):
        col = get_column_letter(start + i)
        # Give enough room for "Mapped: <long ancillary name>" header
        ws.column_dimensions[col].width = max(14, min(40, len(f"Mapped: {src}") + 2))
