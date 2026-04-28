# src/artifacts/excel/tab_data_dictionary.py
"""Generate Data Dictionary tab for Excel CDM."""

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.schema_resolver import (
    ancillary_attribute_index,
    format_ancillary_source_refs,
)
from src.artifacts.common.styles import ExcelStyles


def create_data_dictionary_tab(
    wb: Workbook,
    extractor: CDMExtractor,
    outdir: Optional[Path] = None,
    cdm_name: str = "",
) -> None:
    """
    Create the Data Dictionary tab with all attributes.

    Columns (always present):
      Entity, Attribute, Business Definition, Data Type, Size,
      Nullable, Is PK, Is FK, FK Reference, Classification, PII, PHI, Rematch

    Columns (conditional — only when field code enrichment has been run):
      NCPDP Field Code, EDW F-Code
    """

    ws = wb.create_sheet("Data_Dictionary")

    attributes = extractor.get_all_attributes()

    # Only add field code columns when at least one attribute was enriched.
    # Appears automatically after postprocess_field_codes has run — absent otherwise.
    # No domain checks needed anywhere.
    has_field_codes = any(
        a.ncpdp_field_codes or a.edw_field_codes for a in attributes
    )
    # Detect which ancillary sources have data (one column per source)
    ancillary_sources = set()
    for a in attributes:
        for key in a.source_lineage:
            if key.startswith("ancillary") and a.source_lineage[key]:
                ancillary_sources.add(key)
    ancillary_sources = sorted(ancillary_sources)

    # Per-ancillary attribute indices — recover original schema.table.column
    # from rationalized JSON's source_attribute lists (the rationalizer
    # renames source entities to business-friendly names; this index lets
    # the tab still display the actual PG identifiers).
    ancillary_indices = {}
    if outdir is not None and cdm_name:
        for src in ancillary_sources:
            ancillary_indices[src] = ancillary_attribute_index(outdir, cdm_name, src)

    # --- Headers ---
    headers = [
        "Entity", "Attribute", "Business Definition", "Data Type", "Size",
        "Nullable", "Is PK", "Is FK", "FK Reference",
        "Classification", "PII", "PHI", "Rematch",
    ]
    if has_field_codes:
        headers += ["NCPDP Field Code", "EDW"]
    for anc_src in ancillary_sources:
        display_name = anc_src.replace("ancillary-", "").replace("-", " ").title()
        headers.append(f"Ancillary {display_name}")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        ExcelStyles.apply_header_style(cell)

    # --- Data rows ---
    for row_idx, attr in enumerate(attributes, 2):
        is_alt = row_idx % 2 == 0

        # Size display
        size = ""
        if attr.max_length:
            size = str(attr.max_length)
        elif attr.precision:
            size = f"{attr.precision},{attr.scale or 0}"

        # Rematch flag — combine rematch_type across the attribute's
        # lineage entries.  Legacy entries that only set rematch=True are
        # treated as "S" (semantic) since they predate the name-match pass.
        rematch_types: set = set()
        for mappings in attr.source_lineage.values():
            for mapping in (mappings if isinstance(mappings, list) else []):
                if mapping.get("rematch") is True:
                    rematch_types.add((mapping.get("rematch_type") or "S").upper())
        if not rematch_types:
            rematch_display = ""
        elif rematch_types == {"N"}:
            rematch_display = "N"
        elif rematch_types == {"S"}:
            rematch_display = "S"
        else:
            rematch_display = "B"
        is_rematch = bool(rematch_display)

        row_data = [
            attr.entity_name,
            attr.attribute_name,
            attr.description or "",
            attr.data_type,
            size,
            "Y" if attr.nullable else "N",
            "Y" if attr.pk else "",
            "Y" if attr.fk_to else "",
            attr.fk_to or "",
            attr.classification or "",
            "Y" if attr.is_pii else "",
            "Y" if attr.is_phi else "",
            rematch_display,
        ]

        if has_field_codes:
            row_data += [
                "; ".join(attr.ncpdp_field_codes) if attr.ncpdp_field_codes else "",
                "; ".join(attr.edw_field_codes)   if attr.edw_field_codes   else "",
            ]
        for anc_src in ancillary_sources:
            entries = attr.source_lineage.get(anc_src, [])
            refs = format_ancillary_source_refs(
                ancillary_indices.get(anc_src),
                entries,
            )
            row_data.append("; ".join(refs))

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            ExcelStyles.apply_body_style(cell, is_alt)

            # Highlight PKs and FKs on attribute column
            if col == 2:
                if attr.pk:
                    ExcelStyles.apply_pk_style(cell)
                elif attr.fk_to:
                    ExcelStyles.apply_fk_style(cell)

            # Highlight rematch flag cell in amber
            if col == 13 and is_rematch:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                cell.font = Font(bold=True, color="7D6608")

    # --- Column widths ---
    widths = {
        "A": 25,  # Entity
        "B": 30,  # Attribute
        "C": 50,  # Description
        "D": 15,  # Data Type
        "E": 10,  # Size
        "F": 10,  # Nullable
        "G": 8,   # Is PK
        "H": 8,   # Is FK
        "I": 35,  # FK Reference
        "J": 15,  # Classification
        "K": 8,   # PII
        "L": 8,   # PHI
        "M": 10,  # Rematch
    }
    if has_field_codes:
        widths["N"] = 20  # NCPDP Field Code
        widths["O"] = 20  # EDW F-Code
    # Ancillary column widths (one per source, dynamic position)
    for i, anc_src in enumerate(ancillary_sources):
        col_letter = get_column_letter(len(headers) - len(ancillary_sources) + i + 1)
        widths[col_letter] = 30

    ExcelStyles.set_column_widths(ws, widths)

    # --- Freeze + filter ---
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(attributes) + 1}"