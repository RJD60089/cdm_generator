# src/artifacts/excel/tab_unmapped.py
"""Unmapped Fields tab for Excel CDM."""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.gap_extractor import GapExtractor


def create_unmapped_tab(
    wb: Workbook, 
    extractor: CDMExtractor,
    gap_extractor: GapExtractor = None
) -> Worksheet:
    """
    Create Unmapped Fields tab.
    
    Shows source fields not mapped to CDM from gap analysis.
    """
    
    ws = wb.create_sheet("Unmapped_Fields")
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # Headers
    headers = [
        "Source Type",
        "Source Entity",
        "Source Attribute", 
        "Reason",
        "Suggested CDM Entity",
        "Suggested Attribute Name"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    row = 2
    
    if gap_extractor:
        unmapped = gap_extractor.get_unmapped_fields()
        
        for field in unmapped:
            ws.cell(row=row, column=1, value=field.source_type).border = thin_border
            ws.cell(row=row, column=2, value=field.source_entity).border = thin_border
            ws.cell(row=row, column=3, value=field.source_attribute).border = thin_border
            ws.cell(row=row, column=4, value=field.reason or "").border = thin_border
            ws.cell(row=row, column=4).alignment = wrap_alignment
            ws.cell(row=row, column=5, value=field.suggested_cdm_entity or "").border = thin_border
            ws.cell(row=row, column=6, value=field.suggested_attribute_name or "").border = thin_border
            row += 1
    
    if row == 2:
        # No data
        ws.cell(row=2, column=1, value="No unmapped fields identified")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    
    # Column widths
    widths = [15, 25, 30, 40, 25, 30]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    # Auto-filter
    if row > 2:
        ws.auto_filter.ref = f"A1:F{row-1}"
    
    return ws