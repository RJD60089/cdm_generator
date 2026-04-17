# src/artifacts/excel/tab_business_rules.py
"""Generate Business Rules tab for Excel CDM."""

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.styles import ExcelStyles


CONFLICT_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
CONFLICT_FONT = Font(bold=True, color="C62828")


def _detect_conflicts(rules_full: list) -> list:
    """Identify rules from different sources that may conflict.

    Two rules conflict when they apply to the same attribute, come from
    different sources, and contain contradictory keywords (e.g., one says
    'required' while another says 'optional').
    """
    if len(rules_full) < 2:
        return []

    conflict_pairs = []
    for i, r1 in enumerate(rules_full):
        for r2 in rules_full[i + 1:]:
            s1 = set(r1.get("sources", []))
            s2 = set(r2.get("sources", []))
            if s1 == s2:
                continue
            text1 = r1.get("rule", "").lower()
            text2 = r2.get("rule", "").lower()
            opposites = [
                ("required", "optional"), ("not null", "nullable"),
                ("must", "may"), ("mandatory", "optional"),
            ]
            for a, b in opposites:
                if (a in text1 and b in text2) or (b in text1 and a in text2):
                    conflict_pairs.append((r1, r2))
                    break
    return conflict_pairs


def _summarize_rules(rules: list) -> str:
    """Produce a short summary of multiple rules for one attribute."""
    if not rules:
        return ""
    if len(rules) == 1:
        return rules[0]
    unique = list(dict.fromkeys(rules))
    if len(unique) <= 3:
        return "; ".join(unique)
    return "; ".join(unique[:3]) + f" (+{len(unique) - 3} more)"


def create_business_rules_tab(wb: Workbook, extractor: CDMExtractor) -> None:
    """
    Create the Business Rules tab.

    Columns:
    - Entity, Attribute, Rule Type, Rule Description, Source(s), Conflict
    """

    ws = wb.create_sheet("Business_Rules")

    headers = [
        "Entity", "Attribute", "Rule Type",
        "Rule Description", "Source(s)", "Conflict"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        ExcelStyles.apply_header_style(cell)

    # Collect per-attribute rule groups for conflict detection
    attr_rules = defaultdict(lambda: {"business": [], "validation": []})
    for attr in extractor.get_all_attributes():
        key = (attr.entity_name, attr.attribute_name)
        attr_rules[key]["business"] = attr.business_rules_full
        attr_rules[key]["validation"] = attr.validation_rules_full

    # Detect conflicts per attribute
    attr_conflicts = {}
    for key, groups in attr_rules.items():
        all_rules = groups["business"] + groups["validation"]
        conflicts = _detect_conflicts(all_rules)
        if conflicts:
            conflict_rules = set()
            for r1, r2 in conflicts:
                conflict_rules.add(r1.get("rule", ""))
                conflict_rules.add(r2.get("rule", ""))
            attr_conflicts[key] = conflict_rules
        else:
            attr_conflicts[key] = set()

    row_idx = 2

    for attr in extractor.get_all_attributes():
        key = (attr.entity_name, attr.attribute_name)
        conflicts = attr_conflicts.get(key, set())

        for rule_dict in attr.business_rules_full:
            is_alt = row_idx % 2 == 0
            rule_text = rule_dict.get("rule", "")
            sources = ", ".join(rule_dict.get("sources", []))
            is_conflict = rule_text in conflicts

            row_data = [
                attr.entity_name,
                attr.attribute_name,
                "Business",
                rule_text,
                sources,
                "CONFLICT" if is_conflict else ""
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                ExcelStyles.apply_body_style(cell, is_alt)
                if is_conflict:
                    cell.fill = CONFLICT_FILL
                    if col == 6:
                        cell.font = CONFLICT_FONT

            row_idx += 1

        for rule_dict in attr.validation_rules_full:
            is_alt = row_idx % 2 == 0
            rule_text = rule_dict.get("rule", "")
            sources = ", ".join(rule_dict.get("sources", []))
            is_conflict = rule_text in conflicts

            row_data = [
                attr.entity_name,
                attr.attribute_name,
                "Validation",
                rule_text,
                sources,
                "CONFLICT" if is_conflict else ""
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                ExcelStyles.apply_body_style(cell, is_alt)
                if is_conflict:
                    cell.fill = CONFLICT_FILL
                    if col == 6:
                        cell.font = CONFLICT_FONT

            row_idx += 1

        # Summary row when attribute has multiple rules
        all_rule_texts = attr.business_rules + attr.validation_rules
        if len(all_rule_texts) > 1:
            is_alt = row_idx % 2 == 0
            summary = _summarize_rules(all_rule_texts)

            row_data = [
                attr.entity_name,
                attr.attribute_name,
                "Summary",
                summary,
                "",
                ""
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                ExcelStyles.apply_body_style(cell, is_alt)
                if col == 3:
                    cell.font = Font(bold=True, italic=True)

            row_idx += 1

    # Column widths
    widths = {
        "A": 25,
        "B": 30,
        "C": 15,
        "D": 80,
        "E": 30,
        "F": 12
    }
    ExcelStyles.set_column_widths(ws, widths)

    ws.freeze_panes = "A2"

    if row_idx > 2:
        ws.auto_filter.ref = f"A1:F{row_idx - 1}"
