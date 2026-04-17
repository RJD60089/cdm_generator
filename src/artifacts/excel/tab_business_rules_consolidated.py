# src/artifacts/excel/tab_business_rules_consolidated.py
"""
Generate the AI-consolidated Business Rules tab.

Reads a consolidation JSON produced by
`src/artifacts/common/rule_consolidator.py` and renders per-entity Included
and Rejected rule sets. Conflicts (nullable vs non-nullable, size 5 vs 10,
etc.) are highlighted.

If no consolidation JSON is available, the tab shows a one-row notice so
the workbook still opens cleanly.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from src.artifacts.common.styles import ExcelStyles


INCLUDED_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
REJECTED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
CONFLICT_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
CONFLICT_FONT = Font(bold=True, color="9C0006", size=10)
NOTICE_FONT = Font(italic=True, color="7F7F7F", size=10)


HEADERS = [
    "Entity", "Set", "Attribute", "Consolidated Rule / Rejected Rule",
    "Source Rule IDs", "Sources", "Conflict Type", "Conflict Detail",
    "Rationale / Reason",
]


def _fmt_list(values) -> str:
    if not values:
        return ""
    return ", ".join(str(v) for v in values)


def create_business_rules_consolidated_tab(
    wb: Workbook,
    consolidated_path: Optional[Path],
) -> None:
    """
    Create the Business_Rules_Consolidated tab.

    Args:
        wb: Workbook being assembled.
        consolidated_path: Path to the JSON produced by rule_consolidator.
            If None or missing, renders a placeholder notice.
    """
    ws = wb.create_sheet("Business_Rules_Consolidated")

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        ExcelStyles.apply_header_style(cell)

    if not consolidated_path or not Path(consolidated_path).exists():
        cell = ws.cell(
            row=2, column=1,
            value=(
                "No AI consolidation JSON found. Generate it via the "
                "Business Rules Consolidation option in Step 6 before the "
                "Excel artifact, then regenerate this workbook."
            ),
        )
        cell.font = NOTICE_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
        _set_widths(ws)
        ws.freeze_panes = "A2"
        return

    with open(consolidated_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    row_idx = 2
    for entity in data.get("entities", []):
        entity_name = entity.get("entity_name", "")

        for item in entity.get("included", []):
            is_alt = row_idx % 2 == 0
            conflict_type = (item.get("conflict_type") or "NONE").upper()
            is_conflict = conflict_type not in ("", "NONE")

            row_data = [
                entity_name,
                "Included",
                item.get("attribute_name", ""),
                item.get("consolidated_rule", ""),
                _fmt_list(item.get("source_rule_ids", [])),
                _fmt_list(item.get("sources", [])),
                conflict_type if is_conflict else "",
                item.get("conflict_detail", ""),
                item.get("rationale", ""),
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                ExcelStyles.apply_body_style(cell, is_alt)
                if col == 2:
                    cell.fill = INCLUDED_FILL
                if is_conflict and col in (7, 8):
                    cell.fill = CONFLICT_FILL
                    cell.font = CONFLICT_FONT
            row_idx += 1

        for item in entity.get("rejected", []):
            is_alt = row_idx % 2 == 0
            row_data = [
                entity_name,
                "Rejected",
                item.get("attribute_name", ""),
                item.get("rule", ""),
                str(item.get("source_rule_id", "")) if item.get("source_rule_id") is not None else "",
                _fmt_list(item.get("sources", [])),
                "",
                "",
                item.get("reason", ""),
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                ExcelStyles.apply_body_style(cell, is_alt)
                if col == 2:
                    cell.fill = REJECTED_FILL
            row_idx += 1

    _set_widths(ws)
    ws.freeze_panes = "A2"
    if row_idx > 2:
        last_col = get_column_letter(len(HEADERS))
        ws.auto_filter.ref = f"A1:{last_col}{row_idx - 1}"


def _set_widths(ws) -> None:
    widths = {
        "A": 25,   # Entity
        "B": 11,   # Set
        "C": 28,   # Attribute
        "D": 70,   # Rule text
        "E": 16,   # Source Rule IDs
        "F": 22,   # Sources
        "G": 12,   # Conflict Type
        "H": 45,   # Conflict Detail
        "I": 45,   # Rationale / Reason
    }
    ExcelStyles.set_column_widths(ws, widths)
