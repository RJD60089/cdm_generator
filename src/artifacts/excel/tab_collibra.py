# src/artifacts/excel/tab_collibra.py
"""
Generate the Collibra tab — a flat, Collibra-import-shaped sheet that
mirrors the format expected by the Collibra team for ingestion into the
org's data dictionary.

Each row represents a single CDM attribute mapped to a single physical
source column (Collibra's "[Data Attribute] represents [Column]"
relationship).  Attributes with no source mapping are skipped.

Column layout matches the reference workbook
"All Data Entities & Attributes - Collibra Formatting Structure".
Yellow-highlighted columns are emitted with the yellow fill but no
data — those values are populated by the Collibra team after import.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.lineage_render import (
    EDW_COLUMN_PREFERENCE,
    EDW_TABLE_PREFERENCE,
    first_present,
)
from src.artifacts.common.schema_resolver import SchemaResolver, ancillary_attribute_index
from src.artifacts.common.styles import ExcelStyles
from src.config.config_parser import AppConfig


YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


# Ordered as columns 1..25 in the Collibra reference file.
HEADERS: List[str] = [
    "Full Name",                                                # 1  populated
    "Name",                                                     # 2  populated
    "Asset Id",                                                 # 3  yellow
    "Status",                                                   # 4  yellow
    "Asset Type",                                               # 5  populated (constant)
    "Domain",                                                   # 6  yellow
    "Community",                                                # 7  yellow
    "Domain Type",                                              # 8  yellow
    "Domain Id",                                                # 9  yellow
    "[Data Entity] contains [Data Attribute] > Name",           # 10 yellow
    "[Data Entity] contains [Data Attribute] > Full Name",      # 11 yellow
    "[Data Entity] contains [Data Attribute] > Asset Type",     # 12 yellow
    "[Data Entity] contains [Data Attribute] > Community",      # 13 yellow
    "[Data Entity] contains [Data Attribute] > Domain Type",    # 14 yellow
    "[Data Entity] contains [Data Attribute] > Domain",         # 15 yellow
    "[Data Entity] contains [Data Attribute] > Domain Id",      # 16 yellow
    "[Data Entity] contains [Data Attribute] > Asset Id",       # 17 yellow
    "[Data Attribute] represents [Column] > Name",              # 18 populated
    "[Data Attribute] represents [Column] > Full Name",         # 19 populated
    "[Data Attribute] represents [Column] > Asset Type",        # 20 populated (constant)
    "[Data Attribute] represents [Column] > Community",         # 21 yellow
    "[Data Attribute] represents [Column] > Domain Type",       # 22 yellow
    "[Data Attribute] represents [Column] > Domain",            # 23 yellow
    "[Data Attribute] represents [Column] > Domain Id",         # 24 yellow
    "[Data Attribute] represents [Column] > Asset Id",          # 25 yellow
]

# 1-indexed column numbers that should carry the yellow fill in both
# header and data cells — the Collibra team owns these values.
YELLOW_COL_INDICES = frozenset({3, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 21, 22, 23, 24, 25})

# Per-column widths (1-indexed) tuned to the reference file's content.
COLUMN_WIDTHS: Dict[int, int] = {
    1: 50, 2: 28, 3: 36, 4: 12, 5: 18,
    6: 16, 7: 18, 8: 22, 9: 36,
    10: 38, 11: 50, 12: 18, 13: 16, 14: 22, 15: 18, 16: 36, 17: 36,
    18: 30, 19: 60, 20: 14,
    21: 16, 22: 22, 23: 22, 24: 36, 25: 36,
}


# =============================================================================
# LINEAGE WALKING
# =============================================================================
# Mirrors tab_mapping._lineage_entries — kept local to avoid coupling the two
# tab modules through a private symbol.  If a third tab needs the same logic,
# promote this to src.artifacts.common.

def _lineage_entries(
    attr,
    source_key: str,
    ancillary_index: Optional[Dict[Tuple[str, str], List[Dict[str, str]]]] = None,
) -> List[Dict[str, Any]]:
    """
    Mirror of tab_mapping._lineage_entries — kept local to avoid coupling
    the two tab modules through a private symbol. EDW preference and the
    `first_present` helper are imported from src.artifacts.common.lineage_render
    so the two tabs stay in sync without sharing a private function.
    """
    entries = attr.source_lineage.get(source_key, []) or []
    if isinstance(entries, dict):
        entries = [entries]

    use_ancillary_index = ancillary_index is not None and source_key.startswith("ancillary")
    is_edw = source_key.lower() == "edw"

    out: List[Dict[str, Any]] = []
    for e in entries:
        if not isinstance(e, dict):
            continue
        rationalized_ent = e.get("source_entity", "") or ""
        rationalized_attr = e.get("source_attribute", "") or ""

        if use_ancillary_index:
            originals = ancillary_index.get(
                (rationalized_ent.lower(), rationalized_attr.lower()),
                [],
            )
            if originals:
                for o in originals:
                    out.append({
                        "table":  o.get("table", "")  or rationalized_ent,
                        "column": o.get("column", "") or rationalized_attr,
                        "schema": o.get("schema", ""),
                    })
                continue

        if is_edw:
            edw_table  = first_present(e, EDW_TABLE_PREFERENCE)  or rationalized_ent
            edw_column = first_present(e, EDW_COLUMN_PREFERENCE) or rationalized_attr
            if edw_table or edw_column:
                out.append({
                    "table":  edw_table,
                    "column": edw_column,
                    "schema": "",
                })
            continue

        if rationalized_ent or rationalized_attr:
            out.append({
                "table":  rationalized_ent,
                "column": rationalized_attr,
                "schema": "",
            })
    return out


# =============================================================================
# ROW ASSEMBLY
# =============================================================================

def _collibra_full_name(cdm_name: str, entity: str, attribute: str) -> str:
    """
    Build col 1 'Full Name' = cdm>entity>attribute.

    Underscores in entity/attribute identifiers are preserved verbatim —
    these flow through to S3 columnar storage where spaces are not
    permitted in column names.
    """
    parts = [p for p in (cdm_name, entity, attribute) if p]
    return ">".join(parts)


def _collibra_column_full_name(schema: str, table: str, column: str) -> str:
    """
    Build col 19 '[Data Attribute] represents [Column] > Full Name' as
    schema>table>column(column).  Internal `.` in the schema string is
    preserved (e.g. 'NCRXDB98-Oracle>ncrx098.navitus.local>SQLMGR' stays
    intact) by joining the parts with `>` rather than substituting `.`.
    """
    parts = [p for p in (schema, table, column) if p]
    if not parts:
        return ""
    return ">".join(parts) + "(column)"


def _build_rows(
    extractor: CDMExtractor,
    config: AppConfig,
    schema_resolver: SchemaResolver,
    mapping_sources: List[str],
    ancillary_indices: Dict[str, Dict[Tuple[str, str], List[Dict[str, str]]]],
) -> List[Dict[str, str]]:
    cdm_name = config.cdm.domain or ""
    attributes = sorted(
        extractor.get_all_attributes(),
        key=lambda a: (a.entity_name, a.attribute_name),
    )

    rows: List[Dict[str, str]] = []
    for attr in attributes:
        full_name = _collibra_full_name(cdm_name, attr.entity_name, attr.attribute_name)
        name = attr.attribute_name or ""

        # Dedupe across mapping sources by (schema, table, column) — same
        # physical column referenced from two source keys produces one row.
        merged: Dict[Tuple[str, str, str], Tuple[str, str, str]] = {}
        for src_key in mapping_sources:
            idx = ancillary_indices.get(src_key)
            for entry in _lineage_entries(attr, src_key, ancillary_index=idx):
                src_table = entry["table"]
                src_column = entry["column"]
                schema = (entry.get("schema") or "").strip()
                if not schema and src_table:
                    schema = schema_resolver.resolve(src_key, src_table)
                    if not schema:
                        schema = schema_resolver.resolve("", src_table)
                schema = schema or ""
                key = (schema, src_table, src_column)
                merged.setdefault(key, (schema, src_table, src_column))

        if not merged:
            continue  # Collibra import only wants real source-mapped rows

        for (schema, src_table, src_column) in merged.values():
            rows.append({
                "Full Name": full_name,
                "Name": name,
                "Asset Type": "Data Attribute",
                "[Data Attribute] represents [Column] > Name": src_column,
                "[Data Attribute] represents [Column] > Full Name":
                    _collibra_column_full_name(schema, src_table, src_column),
                "[Data Attribute] represents [Column] > Asset Type": "Column",
            })

    return rows


# =============================================================================
# TAB ENTRY POINT
# =============================================================================

def create_collibra_tab(
    wb: Workbook,
    extractor: CDMExtractor,
    config: AppConfig,
    outdir: Optional[Path] = None,
) -> None:
    """
    Create the Collibra tab — Collibra-formatted asset rows for ingestion.

    Reads ``config.mapping.mapping_sources`` (same lineage keys as the
    Mapping tab) and emits one row per (CDM attribute × source column).
    Columns 3, 4, 6–17, 21–25 are emitted with yellow fill but no data —
    Collibra populates those after import.
    """
    ws = wb.create_sheet("Collibra")

    mapping_cfg = config.mapping
    mapping_sources = list(mapping_cfg.mapping_sources or [])

    if outdir is None:
        outdir = Path(".")
    schema_resolver = SchemaResolver(config, outdir)

    # Header row
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        ExcelStyles.apply_header_style(cell)
        if col_idx in YELLOW_COL_INDICES:
            cell.fill = YELLOW_FILL

    if not mapping_sources:
        cell = ws.cell(
            row=2, column=1,
            value=(
                "No mapping sources configured. Add 'mapping.mapping_sources' "
                "to config.json to populate this tab."
            ),
        )
        cell.font = Font(italic=True, color="7F7F7F", size=10)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
        _apply_widths(ws)
        ws.freeze_panes = "A2"
        return

    # Per-ancillary attribute index (recovers original schema/table/column
    # references for ancillary sources whose entity names were rationalized).
    ancillary_indices: Dict[str, Dict[Tuple[str, str], List[Dict[str, str]]]] = {}
    cdm_name = config.cdm.domain or ""
    for src_key in mapping_sources:
        if src_key.startswith("ancillary"):
            ancillary_indices[src_key] = ancillary_attribute_index(outdir, cdm_name, src_key)

    rows = _build_rows(
        extractor=extractor,
        config=config,
        schema_resolver=schema_resolver,
        mapping_sources=mapping_sources,
        ancillary_indices=ancillary_indices,
    )

    # Write data rows — yellow columns get the fill but no value.
    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(HEADERS, 1):
            value = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = ExcelStyles.BODY_FONT
            cell.alignment = ExcelStyles.BODY_ALIGN
            cell.border = ExcelStyles.THIN_BORDER
            if col_idx in YELLOW_COL_INDICES:
                cell.fill = YELLOW_FILL

    _apply_widths(ws)
    ws.freeze_panes = "A2"
    if rows:
        last_col = get_column_letter(len(HEADERS))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"


def _apply_widths(ws) -> None:
    for idx, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
