# src/artifacts/excel/tab_core_concepts.py
"""Generate Core Concepts tab for Excel CDM."""

from openpyxl import Workbook
from src.config.config_parser import AppConfig
from src.artifacts.common.cdm_extractor import CDMExtractor
from src.artifacts.common.styles import ExcelStyles


def create_core_concepts_tab(
    wb: Workbook,
    extractor: CDMExtractor,
    config: AppConfig
) -> None:
    """
    Create the Core Concepts tab with business definitions.
    
    Columns:
    - Concept Name, Business Definition, CDM Mapping, Key Questions
    """
    
    ws = wb.create_sheet("Core_Concepts")
    
    # Headers
    headers = [
        "Concept Name", "Business Definition", 
        "CDM Mapping", "Key Questions"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        ExcelStyles.apply_header_style(cell)
    
    # Derive concepts from entities
    concepts = _derive_concepts(extractor, config)
    
    # Data rows
    for row_idx, concept in enumerate(concepts, 2):
        is_alt = row_idx % 2 == 0
        
        row_data = [
            concept["name"],
            concept["definition"],
            concept["mapping"],
            concept["questions"]
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            ExcelStyles.apply_body_style(cell, is_alt)
    
    # Column widths
    widths = {
        "A": 25,  # Concept Name
        "B": 60,  # Business Definition
        "C": 30,  # CDM Mapping
        "D": 50   # Key Questions
    }
    ExcelStyles.set_column_widths(ws, widths)
    
    # Freeze header row
    ws.freeze_panes = "A2"


def _derive_concepts(extractor: CDMExtractor, config: AppConfig) -> list:
    """Derive core business concepts from CDM entities."""
    
    concepts = []
    entities = extractor.get_entities()
    
    # Map entity names to business concepts
    for entity in entities:
        # Create human-readable concept name
        concept_name = _entity_to_concept_name(entity.name)
        
        # Use entity description as business definition
        definition = entity.description or f"Represents a {concept_name.lower()} in the system"
        
        # CDM mapping is the entity itself
        mapping = f"{entity.name} entity ({entity.attribute_count} attributes)"
        
        # Generate key questions based on entity
        questions = _generate_questions(entity)
        
        concepts.append({
            "name": concept_name,
            "definition": definition,
            "mapping": mapping,
            "questions": questions
        })
    
    return concepts


def _entity_to_concept_name(entity_name: str) -> str:
    """Convert entity name to human-readable concept."""
    # Handle common patterns
    name = entity_name
    
    # Split camelCase or PascalCase
    import re
    words = re.findall(r'[A-Z][a-z]*|[a-z]+', name)
    
    return " ".join(words)


def _generate_questions(entity) -> str:
    """Generate key business questions for an entity."""
    questions = []
    
    name_lower = entity.name.lower()
    
    # Standard questions based on entity characteristics
    if entity.classification == "Core":
        questions.append(f"What uniquely identifies a {entity.name}?")
    
    if entity.foreign_keys:
        fk_entities = [fk["to_entity"] for fk in entity.foreign_keys]
        questions.append(f"How does this relate to {', '.join(fk_entities[:2])}?")
    
    if "assignment" in name_lower or "association" in name_lower:
        questions.append("What business rules govern this relationship?")
    
    if not questions:
        questions.append(f"What are the valid states for a {entity.name}?")
    
    return " ".join(questions)
