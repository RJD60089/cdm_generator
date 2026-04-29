"""
Guardrails (DGBee Excel) converter.
Converts DGBee format Excel files to JSON.

Sheet filtering precedence (most specific first):
  1. ``include_sheets`` — when provided, only those sheets are kept.
  2. ``exclude_sheets`` — when provided, those sheets are skipped.
  3. Built-in heuristic — case-insensitive matches against a list of
     known noise tabs plus the substring "example" anywhere in the name.

The first filter that matches wins; later ones are not consulted.
"""
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook


# Default heuristic — case-insensitive match. Includes the common typo
# "Glossory" alongside "Glossary".
_DEFAULT_SKIP_NAMES = {
    "glossary", "glossory",
    "dgbee summary", "api summary",
    "data dictionary",
    "pod summary",
    "assumptions", "assumption",
    "not going to the cloud",
    "out of scope", "deprecated",
}
# Substring (case-insensitive) — anything containing one of these is
# treated as template / FHIR-duplicate / supporting content and skipped.
#
# - "example" / "template" — sample data tabs.
# - "fhir" — FHIR reference/value-set/code-system tabs duplicate content
#   already handled by the FHIR rationalizer (which reads FHIR IGs
#   directly).  Including them in guardrails would re-rationalize the
#   same FHIR concepts and inflate the prompt.
_DEFAULT_SKIP_SUBSTRINGS = ("example", "template", "fhir")


def _heuristic_should_skip(sheet_name: str) -> bool:
    n = (sheet_name or "").strip().lower()
    if not n:
        return True
    if n in _DEFAULT_SKIP_NAMES:
        return True
    for sub in _DEFAULT_SKIP_SUBSTRINGS:
        if sub in n:
            return True
    return False


def convert_guardrails_to_json(
    file_path: str,
    include_sheets: Optional[List[str]] = None,
    exclude_sheets: Optional[List[str]] = None,
) -> dict:
    """Convert a DGBee Excel guardrails file to a JSON-friendly dict.

    Args:
        file_path: path to the .xlsx
        include_sheets: when provided, ONLY these tabs are kept
            (case-sensitive exact match by sheet name). Heuristic is not
            consulted.
        exclude_sheets: when provided, these tabs are skipped (exact
            match). Heuristic is consulted for everything else.

    Returns:
        ``{"source_file": <filename>, "sheets": {<name>: [{...rows...}]}}``
    """
    xl = pd.ExcelFile(file_path)
    sheets: Dict[str, Any] = {}

    incl = set(include_sheets or [])
    excl = set(exclude_sheets or [])

    for sheet_name in xl.sheet_names:
        # Precedence 1: explicit include list — strict allow-list mode
        if incl:
            if sheet_name not in incl:
                continue
        else:
            # Precedence 2: explicit exclude list
            if sheet_name in excl:
                continue
            # Precedence 3: built-in heuristic
            if _heuristic_should_skip(sheet_name):
                continue

        df = pd.read_excel(file_path, sheet_name=sheet_name)
        sheets[sheet_name] = df.to_dict('records')

    return {
        'source_file': Path(file_path).name,
        'sheets': sheets,
    }
    
    # Process each "Data Elements" sheet
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('Data Elements'):
            entity_name = sheet_name.replace('Data Elements ', '').strip()
            output["sheets"][entity_name] = _convert_sheet_to_dict(wb[sheet_name])
        elif sheet_name == 'DGBee Summary':
            output["summary"] = _extract_summary(wb[sheet_name])
        elif sheet_name == 'Glossary':
            output["glossary"] = _convert_sheet_to_dict(wb[sheet_name])
    
    return json.dumps(output, indent=2)


def _convert_sheet_to_dict(sheet) -> List[Dict[str, Any]]:
    """
    Convert an Excel sheet to list of dictionaries.
    First row is headers, subsequent rows are data.
    """
    # Get headers from first row (or second row, depending on format)
    headers = []
    header_row = None
    
    # Try to find header row (look for common patterns)
    for row_num in range(1, min(5, sheet.max_row + 1)):
        row = [cell.value for cell in sheet[row_num]]
        # Check if this looks like a header row - FIXED LOGIC
        if any(h and isinstance(h, str) and ('Column' in h or 'Field' in h or 'Name' in h)
               for h in row):
            headers = row
            header_row = row_num
            break
    
    if not headers:
        # Fallback: use first row
        headers = [cell.value for cell in sheet[1]]
        header_row = 1
    
    # Convert data rows to dictionaries
    data = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):  # Skip empty rows
            continue
        
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                # Clean header name
                header = str(headers[i]).strip()
                row_dict[header] = value
        
        if row_dict:  # Only add non-empty rows
            data.append(row_dict)
    
    return data


def _extract_summary(sheet) -> Dict[str, str]:
    """Extract summary information from DGBee Summary sheet"""
    summary = {}
    
    for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
        if row[0] and isinstance(row[0], str):
            # Look for key-value pairs
            if len(row) > 1 and row[1]:
                summary[str(row[0]).strip()] = str(row[1]).strip()
    
    return summary


def extract_entities_from_guardrails(guardrails_json_str: str) -> List[str]:
    """
    Extract entity names from Guardrails JSON string.
    Helper function for analysis.
    
    Args:
        guardrails_json_str: JSON string from convert_guardrails_to_json()
        
    Returns:
        List of entity names
    """
    data = json.loads(guardrails_json_str)
    return list(data.get("sheets", {}).keys())