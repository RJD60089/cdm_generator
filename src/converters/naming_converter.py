"""
Naming standard converter.
Converts enterprise naming standard Excel to JSON rules.
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Any


def convert_naming_standard_to_json(file_path: str) -> str:
    """
    Convert naming standard Excel file to JSON rules.
    
    Args:
        file_path: Path to naming standard Excel file
        
    Returns:
        JSON string representation of naming rules
        
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If Excel file cannot be parsed
    """
    file = Path(file_path)
    
    if not file.exists():
        raise FileNotFoundError(f"Naming standard file not found: {file_path}")
    
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ValueError(f"Failed to load Excel file: {e}")
    
    output = {
        "source_file": file.name,
        "rules": [],
        "conventions": {},
        "data_types": {}
    }
    
    # Process sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Try to extract rules from the sheet
        sheet_data = _extract_naming_rules_from_sheet(ws, sheet_name)
        
        if sheet_data:
            output["rules"].extend(sheet_data.get("rules", []))
            output["conventions"].update(sheet_data.get("conventions", {}))
            output["data_types"].update(sheet_data.get("data_types", {}))
    
    return json.dumps(output, indent=2)


def _extract_naming_rules_from_sheet(sheet, sheet_name: str) -> Dict[str, Any]:
    """
    Extract naming rules from a sheet.
    Tries to identify common patterns in naming standard documents.
    """
    result = {
        "rules": [],
        "conventions": {},
        "data_types": {}
    }
    
    # Get all data from sheet
    headers = None
    data_rows = []
    
    for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if not any(row):  # Skip empty rows
            continue
        
        # Try to identify header row
        if headers is None and any(h and isinstance(h, str) and 
                                   ('standard' in str(h).lower() or 
                                    'rule' in str(h).lower() or
                                    'convention' in str(h).lower() or
                                    'field' in str(h).lower() or
                                    'type' in str(h).lower())
                                         for h in row):
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(row)]
            continue
        
        if headers:
            # Convert row to dict
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value
            data_rows.append(row_dict)
    
    # Process the data based on what we found
    if data_rows:
        for row_dict in data_rows:
            # Try to identify if this is a naming rule
            rule_entry = {
                "sheet": sheet_name,
                "data": row_dict
            }
            result["rules"].append(rule_entry)
    
    return result


def extract_field_conventions(naming_json_str: str) -> Dict[str, str]:
    """
    Extract field naming conventions from naming standard JSON.
    Helper function to get common patterns.
    
    Args:
        naming_json_str: JSON string from convert_naming_standard_to_json()
        
    Returns:
        Dict of field types to naming conventions
    """
    data = json.loads(naming_json_str)
    
    conventions = {}
    
    # Extract common patterns from rules
    for rule in data.get("rules", []):
        rule_data = rule.get("data", {})
        
        # Look for patterns like "date fields end with _dt"
        for key, value in rule_data.items():
            if value and isinstance(value, str):
                key_lower = key.lower()
                if 'suffix' in key_lower or 'convention' in key_lower or 'standard' in key_lower:
                    conventions[key] = value
    
    # Add any explicit conventions
    conventions.update(data.get("conventions", {}))
    
    return conventions
