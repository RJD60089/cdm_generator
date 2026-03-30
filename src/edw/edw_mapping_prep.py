"""
src/edw/edw_mapping_prep.py
============================
EDW Mapping Prep — parses NI_* and NP_* source-to-target XLS files into
structured JSON catalogs for use in CDM generation.

Usage (via orchestrator):
    from src.edw import run_edw_mapping_prep
    run_edw_mapping_prep(initial_dir, persistent_dir, output_dir)

Directory conventions (all relative to project root):
    input/Initial/     — NI_* source-to-target XLS files
    input/Persistent/  — NP_* source-to-target XLS files
    input/edw_catalog/ — output JSON catalog files (one per entity)

    The edw_catalog directory is shared across all CDM domains since entities
    feed multiple CDMs. It is not domain-specific and lives at project root.

File naming convention:
    NI_<ENTITY>_-_source_to_target.xls  (or .xlsx)
    NP_<ENTITY>_-_source_to_target.xls  (or .xlsx)

    Only files containing "source" and "target" in the name are processed.
    Files are matched into NI/NP pairs by entity name extracted from filename.

Functional groups:
    Applied only to entities with 100+ business fields where 30%+ of columns
    match the NCPDP F[1-5]xx naming pattern (e.g. INCYCLEPAID and mirrors).
    Group assignment is structural only — NCPDP cross-reference enrichment
    happens downstream in the EDW rationalizer.
"""

from __future__ import annotations

import json
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import xlrd
except ImportError:
    raise ImportError("xlrd is required: pip install xlrd")

try:
    import openpyxl
except ImportError:
    raise ImportError("openpyxl is required: pip install openpyxl")

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Known SCD2 metadata columns added by NP_ stage
SCD2_COLUMNS = {
    "NP_START_DT",
    "NP_END_DT",
    "NP_PREVIOUS_START_DT",
    "NP_CURRENT_IND",
    "NP_DELETE_IND",
}

# Known ETL-derived columns (not from source system)
DERIVED_COLUMNS = {
    "NI_CHANGE_TYPE_IND",
    "MD5_HASH_KEY",
    "BATCH_ID",
}

# Functional group thresholds
FCODE_PATTERN_THRESHOLD = 0.30   # minimum % of columns matching F[1-5]xx
FCODE_FIELD_COUNT_MIN   = 100    # minimum business field count

# NCPDP Telecom D.0 segment groupings by field number prefix (hundreds digit)
NCPDP_SEGMENT_MAP = {
    range(100, 200): "routing",
    range(200, 300): "pharmacy_id",
    range(300, 400): "member_identity",
    range(400, 500): "claim_submission",
    range(500, 600): "adjudication_response",
}

# CDM domain assignments per entity name.
# Primary domain is first in list.
# Entities shared across CDMs list all applicable domains.
ENTITY_CDM_DOMAINS: dict[str, list[str]] = {
    # Plan & Benefit
    "CARRIER":                      ["plan_and_benefit"],
    "GROUPS":                       ["plan_and_benefit"],
    "SUBGROUPS":                    ["plan_and_benefit"],
    "PLANS":                        ["plan_and_benefit"],
    "GROUPSPLANLIST":               ["plan_and_benefit"],
    "PLANSFMLIST":                  ["plan_and_benefit", "formulary"],
    "PLANSSTLIST":                  ["plan_and_benefit"],
    "PLANDURLIST":                  ["plan_and_benefit", "utilization_management"],
    "PLANCOB":                      ["plan_and_benefit", "eligibility"],
    "PLANCOBLIST":                  ["plan_and_benefit", "eligibility"],
    "CLIENT_HIERARCHY":             ["plan_and_benefit"],
    "GROUPSLOCATIONS":              ["plan_and_benefit"],
    "ASSOCIATION":                  ["plan_and_benefit"],
    "ORGANIZATION":                 ["plan_and_benefit"],
    "HEALTH_PLAN_ACCT_ROLLUP":      ["plan_and_benefit"],
    "ACCTGROUPSLIST":               ["plan_and_benefit"],
    "ACCTMASTER":                   ["plan_and_benefit", "finance"],
    "TPALINK":                      ["plan_and_benefit"],
    # Claims
    "INCYCLEPAID":                  ["claims"],
    "INCYCLEDELETED":               ["claims"],
    "PAIDHISTORY":                  ["claims"],
    "REVHISTORY":                   ["claims"],
    "REJECTS":                      ["claims"],
    "PAIDCHANGELOG":                ["claims"],
    "N1DETAILS":                    ["claims"],
    "SDC":                          ["claims"],
    "BENADJ":                       ["claims"],
    "PAID_LOOPS_SCC_V":             ["claims"],
    "PAID_LOOPS_COMPOUND_V":        ["claims"],
    "PAID_LOOPS_OTHER_PAYER_V":     ["claims"],
    "PDEBASE":                      ["claims"],
    "PDE_CODES":                    ["claims"],
    "CCA_CMO_DATA":                 ["claims"],
    "CCA_PCL_DATA":                 ["claims"],
    # Finance
    "APAR":                         ["finance"],
    "APARBILLTO":                   ["finance"],
    "APARPAYTO":                    ["finance"],
    "APARINVOICESUB":               ["finance"],
    "NET":                          ["finance"],
    "NETLIST":                      ["finance"],
    "PARTDPLANPAY":                 ["finance"],
    "MAC":                          ["finance"],
    # Eligibility
    "SUBSCRIBER":                   ["eligibility"],
    "SUBSCRIBERELIGLIST":           ["eligibility"],
    "SUBSCRIBERGROUPLIST":          ["eligibility", "plan_and_benefit"],
    "SUBSCRIBERACCTLIST":           ["eligibility"],
    "SUBSCRIBERELIGFLAGS":          ["eligibility"],
    "SUBSCRIBEREXCHGROUPLIST":      ["eligibility"],
    "SUBSCRIBEREXCHPLANLIST":       ["eligibility"],
    "SUBSSUBGROUPS":                ["eligibility", "plan_and_benefit"],
    "SUBSCOB":                      ["eligibility"],
    "SUB_ON_DATE":                  ["eligibility"],
    "OLDCARDID":                    ["eligibility"],
    "EXCHANGEREPORTING":            ["eligibility"],
    "MCARECOBCHILD":                ["eligibility"],
    "MCARECOBDTL":                  ["eligibility"],
    "MCARECOBLOG":                  ["eligibility"],
    "MEMPRIORAUTH":                 ["eligibility", "utilization_management"],
    # Drug
    "MASTERDRUG":                   ["drug"],
    "MDDB":                         ["drug"],
    "MDDBTCRF":                     ["drug"],
    "CMS_RXNORM":                   ["drug"],
    "MEDISPAN_CODES":               ["drug"],
    "MED_D_LIMITS":                 ["drug"],
    "COMPOUNDINGREDIENT":           ["drug", "claims"],
    "QUICKCODEGPI":                 ["drug"],
    "QUICKCODES":                   ["drug"],
    "HANDLERAGE":                   ["drug", "plan_and_benefit"],
    "HANDLERCOPAY":                 ["drug", "plan_and_benefit"],
    "HANDLERCOST":                  ["drug", "plan_and_benefit"],
    "HANDLERDISPFEE":               ["drug", "plan_and_benefit"],
    "HANDLERDISPLIMITS":            ["drug", "plan_and_benefit"],
    "HANDLERMISC":                  ["drug", "plan_and_benefit"],
    "HANDLERPF":                    ["drug", "plan_and_benefit"],
    "HANDLERREFILLS":               ["drug", "plan_and_benefit"],
    # Formulary
    "FORMULARYLIST":                ["formulary"],
    "FORMULARYMODIFIER":            ["formulary"],
    "DEPENDENCY":                   ["formulary", "utilization_management"],
    # Prescriber
    "PHYSMASTER":                   ["prescriber"],
    "PHYSNET":                      ["prescriber", "pharmacy"],
    "DEAMASTER":                    ["prescriber"],
    "NUCC_PROVIDER_TAXONOMY_CODE":  ["prescriber"],
    # Pharmacy
    "PHARMACIES":                   ["pharmacy"],
    "NETWORK_ID":                   ["pharmacy", "prescriber"],
    "NET_NETWORK_ID":               ["pharmacy", "prescriber"],
    # Utilization Management
    "PRIORAUTH":                    ["utilization_management"],
    "DURS":                         ["utilization_management"],
    "CUSTOMRULESENGINETRAN":        ["utilization_management"],
    "TRANSITION":                   ["utilization_management"],
    # Reference
    "CODEREF":                      ["reference"],
    "CODEREFTYPES":                 ["reference"],
    "CCA_CODE_DESCRIPTION_XREF":    ["reference"],
}


# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------

def normalise_dtype(raw: str) -> str:
    """Normalise varied data type representations to a consistent form."""
    if not raw:
        return ""
    raw = raw.strip()
    raw = re.sub(r"^V\s*\((\d+)\)$",               r"VARCHAR2(\1)", raw)
    raw = re.sub(r"^VARCHAR2\s*\((\d+)\s*Byte\)$",  r"VARCHAR2(\1)", raw, flags=re.IGNORECASE)
    raw = re.sub(r"^NUM$",                           "NUMBER",        raw, flags=re.IGNORECASE)
    raw = re.sub(r"^DT$",                            "DATE",          raw, flags=re.IGNORECASE)
    raw = re.sub(r"^INT$",                           "INTEGER",       raw, flags=re.IGNORECASE)
    return raw


# Sheet name variants seen across EDW files
_SHEET_NAME_VARIANTS = ["Source_to_Target", "Source to Target", "source_to_target", "source to target", "SourceToTarget", "Sheet1"]


def _find_ni_sheet(sheet_names: list) -> Optional[str]:
    """
    Fallback sheet finder for files with non-standard sheet naming.
    Prefers sheets containing 'NI Target' (multi-target files like PAID_LOOPS_*).
    Falls back to any sheet whose name starts with a known source-to-target prefix.
    Ignores Notes, OVERRIDE, and audit target sheets.
    """
    # Priority 1: sheet explicitly marked as NI Target
    for name in sheet_names:
        if "NI Target" in name or "NI_Target" in name:
            return name
    # Priority 2: any SRC_to_TGT or Source_to_Target sheet (take first)
    for name in sheet_names:
        nl = name.lower()
        if nl.startswith("src_to_tgt") or nl.startswith("source_to_target") or nl.startswith("source to target"):
            return name
    return None


def open_sheet(path: Path, sheet_name: str):
    """
    Open a worksheet from either .xls or .xlsx file.
    Tries exact sheet_name first, then known variants, then fuzzy NI-target fallback.
    Returns (workbook, worksheet, format) where format is 'xls' or 'xlsx'.
    """
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        candidates = [sheet_name] + [v for v in _SHEET_NAME_VARIANTS if v != sheet_name]
        for name in candidates:
            if name in wb.sheetnames:
                return wb, wb[name], "xlsx"
        # Fuzzy fallback for non-standard multi-sheet files
        fallback = _find_ni_sheet(wb.sheetnames)
        if fallback:
            return wb, wb[fallback], "xlsx"
        raise ValueError(f"No sheet named <'{sheet_name}'> in {path.name}. Available: {wb.sheetnames}")

    else:  # .xls
        wb = xlrd.open_workbook(str(path))
        all_names = wb.sheet_names()
        candidates = [sheet_name] + [v for v in _SHEET_NAME_VARIANTS if v != sheet_name]
        for name in candidates:
            try:
                return wb, wb.sheet_by_name(name), "xls"
            except xlrd.biffh.XLRDError:
                continue
        # Fuzzy fallback for non-standard multi-sheet files
        fallback = _find_ni_sheet(all_names)
        if fallback:
            return wb, wb.sheet_by_name(fallback), "xls"
        raise xlrd.biffh.XLRDError(f"No sheet named <'{sheet_name}'>. Available: {all_names}")


def parse_pk(raw) -> Optional[int]:
    """Parse PK order value — may be float (1.0), string, or empty."""
    if raw is None or raw == "":
        return None
    try:
        val = float(str(raw).strip())
        return int(val) if val > 0 else None
    except (ValueError, TypeError):
        return None


def parse_nullable(raw: str) -> bool:
    """Return True if nullable (Y), False if not nullable (N)."""
    return str(raw).strip().upper() != "N"


def extract_notes(wb: xlrd.Book) -> list[str]:
    """Extract meaningful text from the Notes sheet if present."""
    notes = []
    if "Notes" not in wb.sheet_names():
        return notes
    ws = wb.sheet_by_name("Notes")
    seen: set[str] = set()
    for r in range(ws.nrows):
        for c in range(ws.ncols):
            val = str(ws.cell_value(r, c)).strip()
            if val and len(val) > 5 and val not in seen:
                if not any(skip in val for skip in [
                    "Use Initial stage", "Use persistent stage",
                    "See also the Notes", "supply source and target"
                ]):
                    notes.append(val)
                    seen.add(val)
    return notes


def extract_entity_name(filename: str) -> Optional[str]:
    """
    Extract entity name from filename.

    Examples:
        NI_CARRIER - source to target.xls          -> CARRIER
        NP_INCYCLEPAID - source to target.xls      -> INCYCLEPAID
        NI_CLIENT_HIERARCHY - source to target.xls -> CLIENT_HIERARCHY
        NI_MCARECOBCHILD- source to target.xls     -> MCARECOBCHILD
        NP_CODEREFTYPES - source to target.xls     -> CODEREFTYPES
    """
    stem = Path(filename).stem
    stem = re.sub(r"^(NI|NP)_", "", stem, flags=re.IGNORECASE)
    # Strip everything from the separator before "source to target".
    # Handles: " - ", "- ", "_-_", "_", " " etc. followed by "source"
    stem = re.sub(r"[\s_-]+source[\s_-]*to[\s_-]*target.*$", "", stem, flags=re.IGNORECASE)
    stem = stem.strip("_- \t").upper()
    return stem if stem else None


def is_source_to_target_file(path: Path) -> bool:
    """Return True if filename contains both 'source' and 'target'."""
    name = path.name.lower()
    return "source" in name and "target" in name


# ---------------------------------------------------------------------------
# Functional group assignment (F-code entities only)
# ---------------------------------------------------------------------------

def should_use_functional_groups(columns: list[str], business_field_count: int) -> bool:
    """
    Return True only when both conditions are met:
      1. 100+ business fields  — keeps grouping off small tables
      2. 30%+ of columns match NCPDP F[1-5]xx pattern

    The field count floor ensures AI prompts during rationalization receive
    meaningfully-sized chunks rather than over-grouped small tables.
    NCPDP cross-reference enrichment is applied by the rationalizer later.
    """
    if business_field_count < FCODE_FIELD_COUNT_MIN:
        return False
    if not columns:
        return False
    fcode_re = re.compile(r"^F[1-5]\d{2}", re.IGNORECASE)
    matches = sum(1 for c in columns if fcode_re.match(str(c)))
    return (matches / len(columns)) >= FCODE_PATTERN_THRESHOLD


def get_functional_group(source_col: str, ni_col: str) -> Optional[str]:
    """
    Assign functional group from F-code column name.
    Structural assignment only — no external NCPDP lookup required.

    Returns group name or None.
    """
    fcode_re = re.compile(r"^F(\d{3})", re.IGNORECASE)

    for candidate in [source_col or "", ni_col or ""]:
        match = fcode_re.match(candidate)
        if not match:
            match = re.match(r"^F(\d{3})_", candidate, re.IGNORECASE)
        if match:
            num = int(match.group(1))
            for num_range, group in NCPDP_SEGMENT_MAP.items():
                if num in num_range:
                    return group
            return "other_ncpdp"

    # Navitus-specific non-F-code patterns
    upper = (source_col or ni_col or "").upper()
    if any(h in upper for h in ["HANDLER", "HWHY"]):
        return "navitus_handlers"
    if any(c in upper for c in ["CYCLEDATE", "CYCLETIME", "TRANSACTIONTYPE", "ENTEREDDATE"]):
        return "cycle_metadata"

    return None


# ---------------------------------------------------------------------------
# NI file parser
# ---------------------------------------------------------------------------

def _read_sheet_rows(path: Path, sheet_name: str = "Source_to_Target") -> tuple[list[list], object]:
    """
    Open an .xls or .xlsx file and return (rows, workbook) where rows is a
    list-of-lists of string values.  Tries sheet_name variants automatically.
    The workbook is returned so callers can extract Notes if needed.
    """
    wb, ws, fmt = open_sheet(path, sheet_name)

    if fmt == "xlsx":
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v) for v in row])
    else:
        rows = []
        for r in range(ws.nrows):
            rows.append([
                "" if ws.cell_value(r, c) is None else str(ws.cell_value(r, c))
                for c in range(ws.ncols)
            ])
    return rows, wb, fmt


def parse_ni_file(path: Path) -> dict:
    """
    Parse an NI_ source-to-target XLS file.

    Format (13 cols):
        Row 0: Title
        Row 1: Headers (Database-schema | Table | Column | DType | PK | Null? |
                        Table | Column | DType | PK | Null? | Transformation | SR note)
        Row 2+: Data (rows with empty source = Informatica-derived columns)
    """
    rows, wb, fmt = _read_sheet_rows(path)
    notes = extract_notes(wb) if fmt == "xls" else []

    source_db    = None
    source_table = None
    ni_table     = None
    fields: list[dict] = []

    # Locate header row
    header_row = 1
    for r in range(min(3, len(rows))):
        row_vals = [v.strip().lower() for v in rows[r]]
        if "column name" in row_vals:
            header_row = r
            break

    for r in range(header_row + 1, len(rows)):
        row = rows[r]
        ncols = len(row)
        src_db_val  = row[0].strip() if ncols > 0 else ""
        src_tbl_val = row[1].strip() if ncols > 1 else ""
        src_col_val = row[2].strip() if ncols > 2 else ""
        src_dtype   = row[3].strip() if ncols > 3 else ""
        tgt_tbl_val = row[6].strip() if ncols > 6 else ""
        tgt_col_val = row[7].strip() if ncols > 7 else ""
        tgt_dtype   = row[8].strip() if ncols > 8 else ""
        tgt_pk      = row[9]         if ncols > 9 else ""
        tgt_null    = row[10].strip() if ncols > 10 else ""
        transform   = row[11].strip() if ncols > 11 else ""
        sr_note     = row[12].strip() if ncols > 12 else ""

        if not tgt_col_val and not src_col_val:
            continue

        # Capture table/db metadata from first populated rows
        if source_db is None and src_db_val and src_db_val not in ("", "Source"):
            source_db = src_db_val
        if source_table is None and src_tbl_val and src_tbl_val not in ("", "Table Name"):
            source_table = src_tbl_val.upper()
        if ni_table is None and tgt_tbl_val and tgt_tbl_val not in ("", "Table Name"):
            ni_table = tgt_tbl_val.upper()

        is_derived = (
            not src_col_val
            or src_tbl_val in ("", "system date/time", "derived (Y or N)")
            or tgt_col_val.upper() in DERIVED_COLUMNS
            or "derived" in transform.lower()
        )

        transform_clean = transform.replace("\n", " ").strip()
        if sr_note:
            transform_clean = f"{transform_clean} [{sr_note}]".strip(" []")

        fields.append({
            "source_column":       src_col_val.upper() if src_col_val else None,
            "ni_column":           tgt_col_val.upper() if tgt_col_val else None,
            "source_dtype":        normalise_dtype(src_dtype),
            "ni_dtype":            normalise_dtype(tgt_dtype),
            "pk_order_ni":         parse_pk(tgt_pk),
            "nullable_ni":         parse_nullable(tgt_null),
            "transformation_note": transform_clean,
            "is_derived":          is_derived,
        })

    return {
        "source_db":    source_db,
        "source_table": source_table,
        "ni_table":     ni_table,
        "notes":        notes,
        "fields":       fields,
    }


# ---------------------------------------------------------------------------
# NP file parser
# ---------------------------------------------------------------------------

def parse_np_file(path: Path) -> dict:
    """
    Parse an NP_ source-to-target XLS file.

    Format (12 cols):
        Row 0: Title
        Row 1: Headers (Table | Column | DType | PK | Null? |
                        Table | Column | DType | PK | Null? | Transformation | SR note)
        Row 2+: Data (source = NI table columns)
    """
    rows, wb, fmt = _read_sheet_rows(path)
    notes = extract_notes(wb) if fmt == "xls" else []

    ni_table = None
    np_table = None
    fields: list[dict] = []

    # Locate header row
    header_row = 1
    for r in range(min(3, len(rows))):
        row_vals = [v.strip().lower() for v in rows[r]]
        if "column name" in row_vals:
            header_row = r
            break

    for r in range(header_row + 1, len(rows)):
        row = rows[r]
        ncols = len(row)
        src_tbl_val = row[0].strip() if ncols > 0 else ""
        src_col_val = row[1].strip() if ncols > 1 else ""
        tgt_tbl_val = row[5].strip() if ncols > 5 else ""
        tgt_col_val = row[6].strip() if ncols > 6 else ""
        tgt_dtype   = row[7].strip() if ncols > 7 else ""
        tgt_pk      = row[8]         if ncols > 8 else ""
        tgt_null    = row[9].strip() if ncols > 9 else ""
        transform   = row[10].strip() if ncols > 10 else ""
        sr_note     = row[11].strip() if ncols > 11 else ""

        if not tgt_col_val and not src_col_val:
            continue

        if ni_table is None and src_tbl_val and src_tbl_val not in ("", "Table name", "Table Name"):
            ni_table = src_tbl_val.upper()
        if np_table is None and tgt_tbl_val and tgt_tbl_val not in ("", "Table Name"):
            np_table = tgt_tbl_val.upper()

        is_scd2    = tgt_col_val.upper() in SCD2_COLUMNS
        is_derived = (
            not src_col_val
            or src_col_val.lower() in ("system date/time", "derived (y or n)")
            or is_scd2
        )

        transform_clean = transform.replace("\n", " ").strip()
        if sr_note:
            transform_clean = f"{transform_clean} [{sr_note}]".strip(" []")

        fields.append({
            "ni_column":              src_col_val.upper() if src_col_val else None,
            "np_column":              tgt_col_val.upper() if tgt_col_val else None,
            "np_dtype":               normalise_dtype(tgt_dtype),
            "pk_order_np":            parse_pk(tgt_pk),
            "nullable_np":            parse_nullable(tgt_null),
            "transformation_note_np": transform_clean,
            "is_scd2_meta":           is_scd2,
            "is_derived":             is_derived,
        })

    return {
        "ni_table": ni_table,
        "np_table": np_table,
        "notes":    notes,
        "fields":   fields,
    }


# ---------------------------------------------------------------------------
# Merge NI + NP into unified entity catalog entry
# ---------------------------------------------------------------------------

def merge_entity(
    entity_name: str,
    ni_data: Optional[dict],
    np_data: Optional[dict],
) -> dict:
    """
    Merge NI and NP parsed data into a single unified entity catalog entry.

    Matching strategy:
        - NI fields keyed by ni_column
        - NP fields keyed by ni_column (source side of NP mapping)
        - Merged on ni_column as the common key
        - NP-only fields (SCD2, derived) appended at end
    """
    cdm_domains  = ENTITY_CDM_DOMAINS.get(entity_name, [])
    source_db    = ni_data.get("source_db")    if ni_data else None
    source_table = ni_data.get("source_table") if ni_data else None
    ni_table     = (ni_data or {}).get("ni_table") or (np_data or {}).get("ni_table")
    np_table     = (np_data or {}).get("np_table")

    # Deduplicated notes from both files
    all_notes: list[str] = []
    seen_notes: set[str] = set()
    for note in (ni_data or {}).get("notes", []) + (np_data or {}).get("notes", []):
        if note not in seen_notes:
            all_notes.append(note)
            seen_notes.add(note)

    # Index NI fields by ni_column
    ni_fields_by_col: dict[str, dict] = {}
    if ni_data:
        for f in ni_data["fields"]:
            key = f.get("ni_column")
            if key:
                ni_fields_by_col[key] = f

    # Index NP fields by ni_column (source side); collect NP-only (SCD2/derived)
    # A valid ni_column must look like a real column name (uppercase, no spaces)
    # Rows like "system date/time" or "derived (Y or N)" are NP-only derived rows
    _real_col_re = re.compile(r"^[A-Z][A-Z0-9_\[\]]*$")
    np_fields_by_ni_col: dict[str, dict] = {}
    np_only_fields: list[dict] = []
    if np_data:
        for f in np_data["fields"]:
            ni_col = f.get("ni_column")
            if ni_col and _real_col_re.match(ni_col):
                np_fields_by_ni_col[ni_col] = f
            else:
                # Non-column source (system date/time, derived notes) or empty
                np_only_fields.append(f)

    # Union of all NI column keys.
    # NP files reference NI columns as their source — but some NP files reference
    # ETL-derived columns (NI_CHANGE_TYPE_IND, MD5_HASH_KEY) or SCD2 columns
    # (NP_START_DT etc.) that were never real NI business columns.
    # Only add NP-referenced NI cols if they actually exist in ni_fields_by_col,
    # OR if they are not in the known derived/SCD2 sets (i.e. genuinely missing
    # from the NI file due to data entry inconsistency — treat as NP-only).
    all_ni_cols = list(ni_fields_by_col.keys())
    for k in np_fields_by_ni_col:
        if k not in all_ni_cols:
            if k in ni_fields_by_col:
                # Already covered above, but belt-and-suspenders
                all_ni_cols.append(k)
            elif k in SCD2_COLUMNS or k in DERIVED_COLUMNS:
                # Known infrastructure column — move to np_only_fields, not all_ni_cols
                np_only_fields.append(np_fields_by_ni_col[k])
            else:
                # NP references a column the NI file doesn't have — treat as business
                # column that exists in persisted layer only (NP-only)
                np_only_fields.append(np_fields_by_ni_col[k])

    # Pre-count business fields to evaluate functional group thresholds
    provisional_business_count = sum(
        1 for col in all_ni_cols
        if not ni_fields_by_col.get(col, {}).get("is_derived", False)
        and col not in SCD2_COLUMNS
        and col not in DERIVED_COLUMNS
    )

    source_cols = [
        ni_fields_by_col[c].get("source_column") or c
        if c in ni_fields_by_col else c
        for c in all_ni_cols
    ]
    use_functional_groups = should_use_functional_groups(
        source_cols, provisional_business_count
    )

    scd2_columns:    list[str] = []
    derived_columns: list[str] = []
    merged_fields:   list[dict] = []

    for ni_col in all_ni_cols:
        ni_f = ni_fields_by_col.get(ni_col, {})
        np_f = np_fields_by_ni_col.get(ni_col, {})

        source_col = ni_f.get("source_column")
        np_col     = np_f.get("np_column") or ni_col

        # NP dtype is authoritative (persisted form)
        data_type  = np_f.get("np_dtype") or ni_f.get("ni_dtype") or ni_f.get("source_dtype") or ""
        pk_order   = np_f.get("pk_order_np") or ni_f.get("pk_order_ni")
        nullable   = np_f.get("nullable_np") if "nullable_np" in np_f else ni_f.get("nullable_ni", True)
        is_derived = ni_f.get("is_derived", False) or np_f.get("is_derived", False)
        is_scd2    = np_f.get("is_scd2_meta", False) or ni_col in SCD2_COLUMNS

        # Combine transformation notes from both ETL stages
        parts = [
            t for t in [
                ni_f.get("transformation_note", ""),
                np_f.get("transformation_note_np", ""),
            ] if t
        ]
        transform_note = " | ".join(parts)

        if is_scd2:
            scd2_columns.append(np_col)
        if is_derived and ni_col in DERIVED_COLUMNS:
            derived_columns.append(ni_col)

        # cdm_domains: [] for derived/SCD2; entity default for business fields
        field_cdm_domains = [] if (is_derived or is_scd2) else list(cdm_domains)

        field_entry: dict = {
            "source_column":       source_col,
            "ni_column":           ni_col,
            "np_column":           np_col,
            "data_type":           data_type,
            "pk_order":            pk_order,
            "nullable":            nullable,
            "cdm_domains":         field_cdm_domains,
            "transformation_note": transform_note,
            "is_derived":          is_derived,
            "is_scd2_meta":        is_scd2,
        }

        if use_functional_groups:
            field_entry["functional_group"] = get_functional_group(
                source_col or "", ni_col
            )

        merged_fields.append(field_entry)

    # Append NP-only fields (SCD2 metadata, additional derived)
    for np_f in np_only_fields:
        np_col = np_f.get("np_column")
        if not np_col:
            continue

        is_scd2 = np_f.get("is_scd2_meta", False) or np_col in SCD2_COLUMNS
        if is_scd2 and np_col not in scd2_columns:
            scd2_columns.append(np_col)

        # Determine whether this is truly derived/infrastructure or a real business
        # column that simply was not present in the NI file.
        # np_f["is_derived"] is set by the NP parser based on whether src_col_val
        # was empty or a known system phrase — trust that flag.
        ni_col_ref = np_f.get("ni_column")  # May be a column name not in NI file
        is_derived_field = np_f.get("is_derived", True)

        field_entry = {
            "source_column":       None,
            # Preserve the NI column reference even if not in NI file — it tells
            # downstream consumers what the persisted layer was sourced from.
            "ni_column":           ni_col_ref if ni_col_ref and not is_scd2 else None,
            "np_column":           np_col,
            "data_type":           np_f.get("np_dtype", ""),
            "pk_order":            np_f.get("pk_order_np"),
            "nullable":            np_f.get("nullable_np", True),
            "cdm_domains":         [],
            "transformation_note": np_f.get("transformation_note_np", ""),
            "is_derived":          is_derived_field,
            "is_scd2_meta":        is_scd2,
        }

        if use_functional_groups:
            field_entry["functional_group"] = None

        merged_fields.append(field_entry)

    # Build functional_groups summary (business fields only, per group)
    functional_groups_summary = None
    if use_functional_groups:
        from collections import defaultdict
        groups: dict[str, list] = defaultdict(list)
        for f in merged_fields:
            fg = f.get("functional_group")
            sc = f.get("source_column") or f.get("ni_column")
            if fg and sc and not f["is_derived"] and not f["is_scd2_meta"]:
                groups[fg].append(sc)
        functional_groups_summary = dict(groups)

    business_field_count = sum(
        1 for f in merged_fields
        if not f["is_derived"] and not f["is_scd2_meta"]
    )

    entity: dict = {
        "entity_id":             entity_name,
        "source_table":          source_table,
        "source_schema":         "SQLMGR",
        "source_database":       source_db or "PBMCN01",
        "ni_table":              ni_table,
        "np_table":              np_table,
        "cdm_domains":           cdm_domains,
        "mandatory_cdm_mapping": True,
        "processing_notes":      all_notes,
        "scd2_columns":          scd2_columns,
        "derived_columns":       derived_columns,
        "field_count_total":     len(merged_fields),
        "field_count_business":  business_field_count,
        "has_ncpdp_fcodes":      use_functional_groups,
    }

    if functional_groups_summary is not None:
        entity["functional_groups"] = functional_groups_summary

    entity["fields"] = merged_fields
    return entity


# ---------------------------------------------------------------------------
# Main prep class
# ---------------------------------------------------------------------------

class EdwMappingPrep:
    """
    Orchestrates parsing of all NI_/NP_ source-to-target XLS files
    and writes per-entity JSON catalog files to the output directory.
    """

    def __init__(
        self,
        initial_dir:    Path,
        persistent_dir: Path,
        output_dir:     Path,
    ):
        self.initial_dir    = Path(initial_dir)
        self.persistent_dir = Path(persistent_dir)
        self.output_dir     = Path(output_dir)

    def _find_files(self, directory: Path) -> dict[str, Path]:
        """
        Scan directory for source-to-target XLS/XLSX files.
        Returns { ENTITY_NAME_UPPER: path }.
        """
        found: dict[str, Path] = {}
        if not directory.exists():
            logger.warning(f"Directory not found: {directory}")
            return found

        for path in sorted(directory.iterdir()):
            if path.suffix.lower() not in (".xls", ".xlsx"):
                continue
            if not is_source_to_target_file(path):
                continue
            entity = extract_entity_name(path.name)
            if entity:
                found[entity] = path
            else:
                logger.warning(f"Could not extract entity name from: {path.name}")

        logger.info(f"Found {len(found)} source-to-target files in {directory}")
        return found

    def write_catalog_index(self, entity_paths: dict[str, Path]) -> Path:
        """Build and write the EDW catalog index after all entities are parsed.

        The index maps each CDM domain to the list of entity names assigned to it.
        This is consumed by the config generator (Step 0) to populate the ``edw``
        section of each CDM config without scanning the full catalog directory.

        Args:
            entity_paths: Dict of { entity_name: catalog_json_path } as returned
                          by run().

        Returns:
            Path to the written index file.
        """
        from collections import defaultdict

        domain_entities: dict[str, list[str]] = defaultdict(list)
        entity_summaries: list[dict] = []

        for entity_name, entity_file in sorted(entity_paths.items()):
            try:
                data = json.loads(entity_file.read_text(encoding="utf-8"))
            except Exception as e:
                logger.warning(f"Could not read {entity_file.name} for index: {e}")
                continue

            entity_data = data.get("entity", {})
            domains = entity_data.get("cdm_domains", [])

            entity_summaries.append({
                "entity_id":            entity_data.get("entity_id", entity_name),
                "source_table":         entity_data.get("source_table", ""),
                "ni_table":             entity_data.get("ni_table", ""),
                "np_table":             entity_data.get("np_table", ""),
                "cdm_domains":          domains,
                "mandatory_cdm_mapping": entity_data.get("mandatory_cdm_mapping", True),
                "field_count_total":    entity_data.get("field_count_total", 0),
                "field_count_business": entity_data.get("field_count_business", 0),
                "has_ncpdp_fcodes":     entity_data.get("has_ncpdp_fcodes", False),
            })

            for domain in domains:
                domain_entities[domain].append(entity_name)

        # Sort for deterministic output
        entities_by_domain = {
            domain: sorted(names)
            for domain, names in sorted(domain_entities.items())
        }

        index = {
            "_metadata": {
                "generated_by":   "edw_mapping_prep.py",
                "generated_date": datetime.now().isoformat(),
                "total_entities": len(entity_summaries),
                "total_domains":  len(entities_by_domain),
                "notes": [
                    "entities_by_domain: domain → [entity_id, ...] used by config generator",
                    "all_entities: lightweight summary for inspection; full data is in individual catalog files",
                    "Domains match cdm_domains assigned during NI file parse",
                ],
            },
            "entities_by_domain": entities_by_domain,
            "all_entities": sorted(entity_summaries, key=lambda e: e["entity_id"]),
        }

        index_path = self.output_dir / "edw_catalog_index.json"
        index_path.write_text(
            json.dumps(index, indent=2, default=str),
            encoding="utf-8",
        )

        logger.info(f"EDW catalog index written: {index_path}")
        logger.info(
            f"  {len(entity_summaries)} entities across "
            f"{len(entities_by_domain)} CDM domains"
        )
        for domain, names in entities_by_domain.items():
            logger.info(f"    • {domain}: {len(names)} entities")

        return index_path

    def run(self, force: bool = False) -> dict[str, Path]:
        """
        Parse all source-to-target files and write catalog JSON.

        Args:
            force: Re-parse even if output JSON already exists.

        Returns:
            Dict of { entity_name: output_path } for all entities.
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)

        ni_files = self._find_files(self.initial_dir)
        np_files = self._find_files(self.persistent_dir)

        all_entities = sorted(set(list(ni_files.keys()) + list(np_files.keys())))

        logger.info(f"Total unique entities: {len(all_entities)}")
        logger.info(f"  NI+NP pairs: {len(set(ni_files) & set(np_files))}")
        logger.info(f"  NI only:     {len(set(ni_files) - set(np_files))}")
        logger.info(f"  NP only:     {len(set(np_files) - set(ni_files))}")

        processed: dict[str, Path] = {}
        skipped = 0
        errors  = 0

        for entity_name in all_entities:
            output_path = self.output_dir / f"edw_{entity_name.lower()}.json"

            if output_path.exists() and not force:
                logger.info(f"  [SKIP] {entity_name} — catalog exists")
                processed[entity_name] = output_path
                skipped += 1
                continue

            try:
                ni_data = None
                np_data = None

                if entity_name in ni_files:
                    logger.info(f"  [NI]   {ni_files[entity_name].name}")
                    ni_data = parse_ni_file(ni_files[entity_name])

                if entity_name in np_files:
                    logger.info(f"  [NP]   {np_files[entity_name].name}")
                    np_data = parse_np_file(np_files[entity_name])

                entity_catalog = merge_entity(entity_name, ni_data, np_data)

                catalog = {
                    "_metadata": {
                        "entity":         entity_name,
                        "generated_by":   "edw_mapping_prep.py",
                        "generated_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        "ni_source_file": ni_files[entity_name].name if entity_name in ni_files else None,
                        "np_source_file": np_files[entity_name].name if entity_name in np_files else None,
                        "notes": [
                            "cdm_domains at entity level = all CDMs this table contributes to",
                            "cdm_domains at field level = entity default for business fields; [] for derived/SCD2",
                            "mandatory_cdm_mapping: all non-derived, non-SCD2 fields must map to a CDM field",
                            "functional_group present only on entities with 100+ business fields and 30%+ F-code columns",
                            "NCPDP field enrichment (field ID, name, definition) applied in EDW rationalizer",
                        ]
                    },
                    "entity": entity_catalog,
                }

                with open(output_path, "w", encoding="utf-8") as f:
                    json.dump(catalog, f, indent=2, default=str)

                processed[entity_name] = output_path
                logger.info(
                    f"  [OK]   {entity_name} -> {output_path.name} "
                    f"({entity_catalog['field_count_business']} business / "
                    f"{entity_catalog['field_count_total']} total"
                    f"{' / F-code groups' if entity_catalog['has_ncpdp_fcodes'] else ''})"
                )

            except Exception as e:
                logger.error(f"  [ERR]  {entity_name} — {e}", exc_info=True)
                errors += 1

        logger.info(
            f"\nEDW Mapping Prep complete: "
            f"{len(processed) - skipped} processed, "
            f"{skipped} skipped, "
            f"{errors} errors"
        )

        # Write catalog index for config generator consumption
        if processed:
            self.write_catalog_index(processed)

        return processed


# ---------------------------------------------------------------------------
# Orchestrator entry point
# ---------------------------------------------------------------------------

def run_edw_mapping_prep(
    initial_dir:    str | Path,
    persistent_dir: str | Path,
    output_dir:     str | Path,
    force:          bool = False,
    dry_run:        bool = False,
) -> dict[str, Path]:
    """
    Entry point called from the CDM orchestrator.

    Args:
        initial_dir:    Path to input/Initial/ (NI_ XLS files)
        persistent_dir: Path to input/Persistent/ (NP_ XLS files)
        output_dir:     Path to input/edw_catalog/ (JSON output)
        force:          Re-parse even if catalog JSON already exists
        dry_run:        Scan and report only — do not write files

    Returns:
        Dict of { entity_name: output_path }
    """
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S",
    )

    print(f"\n{'='*60}")
    print(f"EDW MAPPING PREP")
    print(f"{'='*60}")
    print(f"  Initial dir:    {initial_dir}")
    print(f"  Persistent dir: {persistent_dir}")
    print(f"  Output dir:     {output_dir}")
    print(f"  Mode:           {'DRY RUN' if dry_run else 'LIVE'}")
    print(f"  Force reparse:  {force}")

    if dry_run:
        prep = EdwMappingPrep(initial_dir, persistent_dir, output_dir)
        ni_files     = prep._find_files(prep.initial_dir)
        np_files     = prep._find_files(prep.persistent_dir)
        all_entities = sorted(set(list(ni_files.keys()) + list(np_files.keys())))

        print(f"\n  DRY RUN — {len(all_entities)} entities found:")
        unmapped = []
        for entity in all_entities:
            ni_mark    = "NI✓" if entity in ni_files else "NI✗"
            np_mark    = "NP✓" if entity in np_files else "NP✗"
            domains    = ENTITY_CDM_DOMAINS.get(entity)
            domain_str = ", ".join(domains) if domains else "⚠ UNMAPPED"
            if not domains:
                unmapped.append(entity)
            print(f"    {entity:<45} [{ni_mark}] [{np_mark}]  {domain_str}")

        if unmapped:
            print(f"\n  ⚠  No CDM domain assigned for: {unmapped}")
            print(f"     Add entries to ENTITY_CDM_DOMAINS in edw_mapping_prep.py")

        return {}

    prep = EdwMappingPrep(initial_dir, persistent_dir, output_dir)
    return prep.run(force=force)