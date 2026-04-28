# src/cdm_full/postprocess_field_codes.py
"""
Post-processing: NCPDP / EDW Field Code Enrichment

Adds two new fields to every CDM attribute that has relevant source lineage:

  ncpdp_field_codes : list[str]  e.g. ["512-FC", "369-2Q"]
      Standard NCPDP Telecom field identifiers, sourced from the rationalized
      NCPDP file's source_metadata.source_ref values.

  edw_field_codes   : list[str]  e.g. ["F201", "F462", "UN002"]
      EDW field codes extracted from EDW lineage source_attribute values
      that match the pattern F/UN/UT + digits (raw NCPDP codes carried in
      EDW source-to-target mappings).

These are stored in the CDM JSON so they flow through to any downstream
artifact (Excel, Word doc, DDL comments, etc.).

Most valuable for Claims CDMs where NCPDP field numbers are meaningful to
analysts.  Can be run for any domain — if the rationalized NCPDP source is
not present, NCPDP codes are skipped and only EDW F-codes are extracted.

No LLM calls — purely structural enrichment.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Any

# ---------------------------------------------------------------------------
# F-code / UN / UT pattern that appears in EDW source_attribute values
# ---------------------------------------------------------------------------
_FIELD_CODE_RE = re.compile(
    r"^(?:F\d{3}|UN\d{3}|UT\d{3}|[A-Z]\d{2}-[A-Z]{2})$",
    re.IGNORECASE
)


# ---------------------------------------------------------------------------
# File finders
# ---------------------------------------------------------------------------

def _find_rationalized(outdir: Path, domain: str, source_type: str) -> Optional[Path]:
    """Find latest rationalized file for a given source type."""
    rat_dir = outdir / "rationalized"
    if not rat_dir.exists():
        return None
    domain_safe = domain.lower().replace(" ", "_")
    matches = sorted(
        rat_dir.glob(f"rationalized_{source_type}_{domain_safe}_*.json"),
        reverse=True
    )
    return matches[0] if matches else None


def _find_latest_excel(outdir: Path, domain: str) -> Optional[Path]:
    """Find the latest Excel workbook in outdir/artifacts/."""
    artifacts_dir = outdir / "artifacts"
    if not artifacts_dir.exists():
        return None
    domain_safe = domain.lower().replace(" ", "_")
    matches = sorted(
        artifacts_dir.glob(f"{domain_safe}_CDM_*.xlsx"),
        reverse=True
    )
    return matches[0] if matches else None


# ---------------------------------------------------------------------------
# Build NCPDP field-code lookup
# ---------------------------------------------------------------------------

def _build_ncpdp_lookup(ncpdp_path: Path) -> Dict[str, str]:
    """
    Build lookup: rationalized_attr_name -> NCPDP field code string.

    Source: rationalized NCPDP JSON, entities[*].attributes[*].source_metadata.source_ref
    source_ref format: "512-FC | T"  ->  field code = "512-FC"
    """
    lookup: Dict[str, str] = {}
    with open(ncpdp_path, "r", encoding="utf-8") as f:
        ncpdp = json.load(f)

    for entity in ncpdp.get("entities", []):
        for attr in entity.get("attributes", []):
            attr_name = attr.get("attribute_name", "")
            if not attr_name:
                continue
            source_ref = (attr.get("source_metadata") or {}).get("source_ref", "")
            if source_ref:
                # Take the part before the first "|"
                field_code = source_ref.split("|")[0].strip()
                if field_code:
                    lookup[attr_name] = field_code

    return lookup


# ---------------------------------------------------------------------------
# Enrichment
# ---------------------------------------------------------------------------

def _enrich_cdm(
    cdm: Dict[str, Any],
    ncpdp_lookup: Dict[str, str]
) -> tuple[Dict[str, Any], int, int]:
    """
    Walk CDM attributes and add ncpdp_field_codes / edw_field_codes.

    Returns:
        updated cdm, attrs_with_ncpdp_codes, attrs_with_edw_codes
    """
    ncpdp_count = edw_count = 0

    for entity in cdm.get("entities", []):
        for attr in entity.get("attributes", []):
            lineage = attr.get("source_lineage", {})

            # --- NCPDP field codes ---
            ncpdp_codes: List[str] = []
            for entry in lineage.get("ncpdp", []):
                src_attr = entry.get("source_attribute", "")
                code = ncpdp_lookup.get(src_attr)
                if code and code not in ncpdp_codes:
                    ncpdp_codes.append(code)

            if ncpdp_codes:
                attr["ncpdp_field_codes"] = ncpdp_codes
                ncpdp_count += 1

            # --- EDW field codes ---
            edw_codes: List[str] = []
            for entry in lineage.get("edw", []):
                src_entity = entry.get("source_entity", "")
                src_attr = entry.get("source_attribute", "")
                if src_attr:
                    ref = f"{src_entity}.{src_attr}" if src_entity else src_attr
                    if ref not in edw_codes:
                        edw_codes.append(ref)

            if edw_codes:
                attr["edw_field_codes"] = edw_codes
                edw_count += 1

    return cdm, ncpdp_count, edw_count


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run_field_codes_postprocess(
    cdm: Dict[str, Any],
    llm: Any,                   # Not used — kept for registry signature compat
    dry_run: bool = False,
    gaps_path: Optional[Path] = None,
    outdir: Optional[Path] = None,
    domain: str = ""
) -> Dict[str, Any]:
    """
    Enrich CDM attributes with NCPDP and EDW field codes.

    Args:
        cdm:      Full CDM dictionary (modified in place)
        llm:      Unused — present for registry interface compatibility
        dry_run:  If True, show what would be done without modifying
        outdir:   Base output directory (to find rationalized/ folder)
        domain:   CDM domain name (used to locate rationalized files)

    Returns:
        Updated CDM dictionary
    """
    print(f"\n   POST-PROCESSING: NCPDP / EDW Field Code Enrichment")
    print(f"   {'-'*40}")

    # Locate rationalized NCPDP file
    ncpdp_path = None
    if outdir:
        ncpdp_path = _find_rationalized(outdir, domain, "ncpdp")

    if not ncpdp_path:
        print(f"   ⚠️  Rationalized NCPDP file not found in {outdir}/rationalized/")
        print(f"       EDW F-codes will still be extracted from lineage directly.")
        ncpdp_lookup: Dict[str, str] = {}
    else:
        print(f"   NCPDP source : {ncpdp_path.name}")
        ncpdp_lookup = _build_ncpdp_lookup(ncpdp_path)
        print(f"   NCPDP lookup : {len(ncpdp_lookup)} attrs indexed")

    if dry_run:
        print(f"\n   DRY RUN — no changes applied")
        # Show sample of what would be added
        sample_ncpdp = list(ncpdp_lookup.items())[:5]
        print(f"\n   Sample NCPDP codes:")
        for k, v in sample_ncpdp:
            print(f"      {k} → {v}")
        return cdm

    # Enrich
    cdm, ncpdp_count, edw_count = _enrich_cdm(cdm, ncpdp_lookup)

    total_attrs = sum(len(e.get("attributes", [])) for e in cdm.get("entities", []))
    print(f"   Total CDM attributes  : {total_attrs}")
    print(f"   With NCPDP field codes: {ncpdp_count}")
    print(f"   With EDW field codes  : {edw_count}")

    # Sample output
    print(f"\n   Sample enriched attributes:")
    shown = 0
    for entity in cdm.get("entities", []):
        for attr in entity.get("attributes", []):
            ncpdp_codes = attr.get("ncpdp_field_codes", [])
            edw_codes   = attr.get("edw_field_codes", [])
            if ncpdp_codes or edw_codes:
                ncpdp_str = ", ".join(ncpdp_codes) if ncpdp_codes else "—"
                edw_str   = ", ".join(edw_codes)   if edw_codes   else "—"
                print(f"      {entity['entity_name']}.{attr['attribute_name']}")
                print(f"        NCPDP: {ncpdp_str}")
                print(f"        EDW  : {edw_str}")
                shown += 1
                if shown >= 5:
                    break
        if shown >= 5:
            break

    # Update Excel workbook in-place
    if outdir:
        _update_excel_data_dictionary(cdm, outdir, domain)

    return cdm


# ---------------------------------------------------------------------------
# In-place Excel Data_Dictionary tab replacement
# ---------------------------------------------------------------------------

def _update_excel_data_dictionary(
    cdm: Dict[str, Any],
    outdir: Path,
    domain: str
) -> None:
    """
    Replace the Data_Dictionary sheet in the latest Excel workbook with a
    rebuilt version that includes NCPDP and EDW field code columns.

    Opens the workbook, removes the existing Data_Dictionary sheet, inserts
    a fresh one at the same position, then saves back to the same file.
    All other tabs are untouched.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(f"   ⚠️  openpyxl not available — skipping Excel update")
        return

    xlsx_path = _find_latest_excel(outdir, domain)
    if not xlsx_path:
        print(f"\n   ⚠️  No Excel workbook found in {outdir / 'artifacts'} — skipping tab update")
        print(f"       Run Step 7 first, then re-run this post-process step.")
        return

    print(f"\n   Updating Excel workbook: {xlsx_path.name}")

    wb = load_workbook(xlsx_path)
    sheet_names = wb.sheetnames

    # Find current position of Data_Dictionary
    if "Data_Dictionary" not in sheet_names:
        print(f"   ⚠️  Data_Dictionary tab not found in workbook — skipping")
        wb.close()
        return

    tab_position = sheet_names.index("Data_Dictionary")

    # Remove existing sheet
    del wb["Data_Dictionary"]

    # Rebuild it using the standard tab generator (which reads field codes
    # from AttributeDetail.ncpdp_field_codes / edw_field_codes)
    try:
        from src.artifacts.common.cdm_extractor import CDMExtractor
        from src.artifacts.excel.tab_data_dictionary import create_data_dictionary_tab

        # Write enriched CDM to a temp path for the extractor to read
        import tempfile, json as _json
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False, encoding="utf-8"
        ) as tmp:
            _json.dump(cdm, tmp)
            tmp_path = Path(tmp.name)

        extractor = CDMExtractor(cdm_path=tmp_path)
        # outdir + cdm_name let the tab look up rationalized JSON to render
        # original schema.table.column refs in ancillary columns.
        create_data_dictionary_tab(wb, extractor, outdir=outdir, cdm_name=domain)
        tmp_path.unlink(missing_ok=True)

        # Move the new sheet to the original position
        wb.move_sheet("Data_Dictionary", offset=tab_position - len(wb.sheetnames) + 1)

        wb.save(xlsx_path)
        print(f"   ✓ Data_Dictionary tab replaced at position {tab_position + 1}")
        print(f"     NCPDP Field Code and EDW F-Code columns added")

    except Exception as e:
        print(f"   ⚠️  Excel update failed: {e}")
        print(f"       Re-run Step 7 to generate a fresh workbook with field codes.")
    finally:
        try:
            wb.close()
        except Exception:
            pass