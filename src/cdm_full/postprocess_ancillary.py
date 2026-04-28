# src/cdm_full/postprocess_ancillary.py
"""
Post-processing: Ancillary Source Enrichment

Adds ancillary source references to every CDM attribute that has
ancillary source lineage entries. Follows the exact pattern of
postprocess_field_codes.py.

For each attribute with source_lineage["ancillary"] entries, extracts
the source reference (schema.table.column) and stores it as
attr["ancillary_source_refs"] — a list of strings.

These are stored in the CDM JSON so they flow through to downstream
artifacts (Excel Data Dictionary, Word doc, etc.).

No LLM calls — purely structural enrichment.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Dict, List, Optional, Any


# ---------------------------------------------------------------------------
# File finders (same pattern as postprocess_field_codes.py)
# ---------------------------------------------------------------------------

def _find_rationalized_ancillary_files(outdir: Path, domain: str) -> List[Path]:
    """Find all rationalized ancillary files (any source_id)."""
    rat_dir = outdir / "rationalized"
    if not rat_dir.exists():
        return []
    domain_safe = domain.lower().replace(" ", "_")
    # Match both old format (rationalized_ancillary_{domain}_*)
    # and new format (rationalized_ancillary-{id}_{domain}_*)
    matches = sorted(
        [f for f in rat_dir.glob(f"rationalized_ancillary*_{domain_safe}_*.json")],
        reverse=True,
    )
    return matches


def _find_latest_excel(outdir: Path, domain: str) -> Optional[Path]:
    """Find the latest Excel workbook in outdir/artifacts/."""
    artifacts_dir = outdir / "artifacts"
    if not artifacts_dir.exists():
        return None
    domain_safe = domain.lower().replace(" ", "_")
    matches = sorted(
        artifacts_dir.glob(f"{domain_safe}_CDM_*.xlsx"),
        reverse=True,
    )
    return matches[0] if matches else None


# ---------------------------------------------------------------------------
# Build ancillary lookup
# ---------------------------------------------------------------------------

def _build_ancillary_lookup(ancillary_path: Path) -> Dict[str, str]:
    """
    Build lookup: rationalized_attr_name -> ancillary source reference string.

    Source: rationalized ancillary JSON,
    entities[*].attributes[*].source_attribute (list of source refs)

    Returns dict mapping attribute_name -> semicolon-joined source refs.
    """
    lookup: Dict[str, str] = {}
    with open(ancillary_path, "r", encoding="utf-8") as f:
        ancillary = json.load(f)

    for entity in ancillary.get("entities", []):
        entity_name = entity.get("entity_name", "")
        for attr in entity.get("attributes", []):
            attr_name = attr.get("attribute_name", "")
            if not attr_name:
                continue
            source_refs = attr.get("source_attribute", attr.get("source_files", []))
            if source_refs:
                # Key by attribute name; store all source refs
                key = attr_name.lower()
                ref_str = "; ".join(source_refs) if isinstance(source_refs, list) else str(source_refs)
                lookup[key] = ref_str

    return lookup


# ---------------------------------------------------------------------------
# Enrichment
# ---------------------------------------------------------------------------

def _enrich_cdm(
    cdm: Dict[str, Any],
    ancillary_lookup: Dict[str, str],
) -> tuple:
    """
    Walk CDM attributes and add ancillary_source_refs.

    Handles multiple independent ancillary sources — iterates all
    source_lineage keys starting with 'ancillary'.

    Returns:
        (updated cdm, attrs_with_ancillary_refs count)
    """
    anc_count = 0

    for entity in cdm.get("entities", []):
        for attr in entity.get("attributes", []):
            lineage = attr.get("source_lineage", {})

            ancillary_refs: List[str] = []
            for key, entries in lineage.items():
                if not key.startswith("ancillary"):
                    continue
                if not isinstance(entries, list):
                    continue
                for entry in entries:
                    src_entity = entry.get("source_entity", "")
                    src_attr = entry.get("source_attribute", "")
                    if src_entity and src_attr:
                        ancillary_refs.append(f"{key}:{src_entity}.{src_attr}")
                    elif src_attr:
                        ancillary_refs.append(f"{key}:{src_attr}")

            if ancillary_refs:
                attr["ancillary_source_refs"] = ancillary_refs
                anc_count += 1

    return cdm, anc_count


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run_ancillary_postprocess(
    cdm: Dict[str, Any],
    llm: Any,                   # Not used — kept for registry signature compat
    dry_run: bool = False,
    gaps_path: Optional[Path] = None,
    outdir: Optional[Path] = None,
    domain: str = "",
) -> Dict[str, Any]:
    """
    Enrich CDM attributes with ancillary source references.

    Args:
        cdm:      Full CDM dictionary (modified in place)
        llm:      Unused — present for registry interface compatibility
        dry_run:  If True, show what would be done without modifying
        outdir:   Base output directory (to find rationalized/ folder)
        domain:   CDM domain name (used to locate rationalized files)

    Returns:
        Updated CDM dictionary
    """
    print(f"\n   POST-PROCESSING: Ancillary Source Enrichment")
    print(f"   {'-' * 40}")

    # Locate rationalized ancillary files (may be multiple with independent source_ids)
    ancillary_paths = []
    if outdir:
        ancillary_paths = _find_rationalized_ancillary_files(outdir, domain)

    if not ancillary_paths:
        print(f"   No rationalized ancillary file found — skipping")
        return cdm

    # Build combined lookup from all ancillary files
    ancillary_lookup: Dict[str, str] = {}
    for ancillary_path in ancillary_paths:
        print(f"   Ancillary source: {ancillary_path.name}")
        file_lookup = _build_ancillary_lookup(ancillary_path)
        ancillary_lookup.update(file_lookup)
    print(f"   Ancillary lookup: {len(ancillary_lookup)} attrs indexed")

    if dry_run:
        print(f"\n   DRY RUN — no changes applied")
        sample = list(ancillary_lookup.items())[:5]
        print(f"\n   Sample ancillary refs:")
        for k, v in sample:
            print(f"      {k} -> {v}")
        return cdm

    # Enrich
    cdm, anc_count = _enrich_cdm(cdm, ancillary_lookup)

    total_attrs = sum(len(e.get("attributes", [])) for e in cdm.get("entities", []))
    print(f"   Total CDM attributes      : {total_attrs}")
    print(f"   With ancillary source refs : {anc_count}")

    # Sample output
    print(f"\n   Sample enriched attributes:")
    shown = 0
    for entity in cdm.get("entities", []):
        for attr in entity.get("attributes", []):
            refs = attr.get("ancillary_source_refs", [])
            if refs:
                print(f"      {entity['entity_name']}.{attr['attribute_name']}")
                print(f"        Ancillary: {'; '.join(refs)}")
                shown += 1
                if shown >= 5:
                    break
        if shown >= 5:
            break

    # Update Excel workbook in-place
    if outdir:
        _update_excel_data_dictionary(cdm, outdir, domain)

    return cdm


# ---------------------------------------------------------------------------
# In-place Excel Data_Dictionary tab replacement
# ---------------------------------------------------------------------------

def _update_excel_data_dictionary(
    cdm: Dict[str, Any],
    outdir: Path,
    domain: str,
) -> None:
    """
    Replace the Data_Dictionary sheet in the latest Excel workbook with a
    rebuilt version that includes ancillary source columns.

    Follows the exact same pattern as postprocess_field_codes.py lines 239-309.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(f"   openpyxl not available — skipping Excel update")
        return

    xlsx_path = _find_latest_excel(outdir, domain)
    if not xlsx_path:
        print(f"\n   No Excel workbook found in {outdir / 'artifacts'} — skipping tab update")
        print(f"       Run Step 7 first, then re-run this post-process step.")
        return

    print(f"\n   Updating Excel workbook: {xlsx_path.name}")

    wb = load_workbook(xlsx_path)
    sheet_names = wb.sheetnames

    if "Data_Dictionary" not in sheet_names:
        print(f"   Data_Dictionary tab not found in workbook — skipping")
        wb.close()
        return

    tab_position = sheet_names.index("Data_Dictionary")

    # Remove existing sheet
    del wb["Data_Dictionary"]

    try:
        from src.artifacts.common.cdm_extractor import CDMExtractor
        from src.artifacts.excel.tab_data_dictionary import create_data_dictionary_tab

        import tempfile
        import json as _json

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False, encoding="utf-8"
        ) as tmp:
            _json.dump(cdm, tmp)
            tmp_path = Path(tmp.name)

        extractor = CDMExtractor(cdm_path=tmp_path)
        # outdir + domain let the tab look up rationalized JSON for
        # ancillary columns to render original schema.table.column refs.
        create_data_dictionary_tab(wb, extractor, outdir=outdir, cdm_name=domain)
        tmp_path.unlink(missing_ok=True)

        wb.move_sheet("Data_Dictionary", offset=tab_position - len(wb.sheetnames) + 1)

        wb.save(xlsx_path)
        print(f"   Data_Dictionary tab replaced at position {tab_position + 1}")
        print(f"     Ancillary Source column added")

    except Exception as e:
        print(f"   Excel update failed: {e}")
        print(f"       Re-run Step 7 to generate a fresh workbook.")
    finally:
        try:
            wb.close()
        except Exception:
            pass
